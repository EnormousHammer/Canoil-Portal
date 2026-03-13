"""
Price List Service for Canoil Portal.

Reads "Copy of Price List Master 2026.xlsx" from Google Drive using the
service account (canoil-backend@dulcet-order-474521-q1.iam.gserviceaccount.com),
parses all sheets, and provides a clean search interface for the AI chat.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
FILE STRUCTURE (5 sheets)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Sheet               Purpose
──────────────────  ──────────────────────────────────────────────────────────
Sept 2024 prices    MASTER — all customers & products, full price history.
                    Has TWO explicit current-price columns (most authoritative):
                      col 21 → "Current Grease Price as of Oct '24 increase"
                      col 22 → "Current Reolube Price as of Apr '25 increase"
                    NOTE: Apr '25 Reolube is THE MOST RECENT price of the whole file.

Mail merge Reolube  Reolube subset — current price = col 5 (Sept 2024).
                    OLDER than master's Apr 2025 Reolube column. Used only as fallback.

Mail merge Grease   Grease subset — current price = col 6 (Sept 2024 Grease).
                    Roughly same as master's Oct 2024 Grease column.

Mail Merge MOV VSG  MOV / VSG subset — current price = col 5 (Sept 2024 / Oct 2024).
                    Roughly same as master's Oct 2024 Grease column.

sort                Duplicate of master, sorted differently. Skipped (master preferred).

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
KEY COLUMNS (every sheet)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  Customer                  – company name (customer-specific price list)
  Product Description/Code  – product name / code
  Product Size              – packaging: Drum 180kg, Pail (17kg), Case (30),
                              Keg (55kg), 3x10 tube case, Tote, etc.
  Currency                  – CAD or USD (per customer agreement)
  <price columns>           – historical, then EXPLICIT current:
                              "Current Grease Price as of Oct '24 increase"  (grease/MOV/VSG)
                              "Current Reolube Price as of Apr '25 increase" (reolube)

Cached for 30 minutes.
"""

import re
import time
from io import BytesIO
from typing import Optional, Dict, Any, List

# ── Constants ─────────────────────────────────────────────────────────────────
PRICE_LIST_FILE_NAME = "Copy of Price List Master 2026"
CACHE_TTL_SECONDS    = 1800   # 30 minutes

# Definitively "current" column names (substring match, lowercase).
# Checked first in _current_price(); the first match wins.
# Order matters: Reolube is checked before Grease because the Reolube column
# is Apr 2025 (most recent) and some rows have both columns present.
CURRENT_COL_KEYWORDS = [
    ("current reolube price", "Current Reolube Price"),
    ("current grease price",  "Current Grease Price"),
]

# Columns that are never prices
NON_PRICE_COLS = {
    "contact information", "contact email", "customer",
    "product description /code", "product description/code",
    "product description / code", "product size", "currency", "old price",
}

MASTER_SHEET = "Sept 2024 prices"

# ── Module-level cache ────────────────────────────────────────────────────────
_cache: Dict[str, Any] = {}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _get_drive_service():
    try:
        from google_drive_service import GoogleDriveService
        svc = GoogleDriveService()
        if svc.authenticate():
            return svc
    except Exception as exc:
        print(f"[PriceList] Could not get Drive service: {exc}")
    return None


def _search_file(gdrive_svc) -> Optional[Dict[str, str]]:
    """Search for the price list file across all drives the service account can see."""
    try:
        service = gdrive_svc._get_fresh_service()
        query = "name contains 'Price List Master 2026' and trashed=false"
        results = service.files().list(
            q=query,
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
            fields="files(id, name, modifiedTime)",
            orderBy="modifiedTime desc",
            pageSize=10,
        ).execute()
        files = results.get("files", [])
        if files:
            chosen = files[0]
            print(f"[PriceList] Found: '{chosen['name']}' (id={chosen['id']})")
            return {"id": chosen["id"], "name": chosen["name"]}
        print("[PriceList] Price list file not found on Google Drive.")
        return None
    except Exception as exc:
        print(f"[PriceList] File search failed: {exc}")
        return None


def _normalize(s: str) -> str:
    """Collapse whitespace and lowercase."""
    return re.sub(r"\s+", " ", str(s)).strip().lower()


def _is_price_col(col_name: str) -> bool:
    """True if the column header is a date/price column (not an ID/name/currency column)."""
    if col_name is None:
        return False
    clean = _normalize(col_name)
    for skip in NON_PRICE_COLS:
        if clean.startswith(skip):
            return False
    # Pure multiplier header (e.g. "1.06", "6%")
    if re.match(r"^[\d\.%]+$", clean):
        return False
    return True


def _current_price(row_orig: dict, headers: List[str]) -> tuple:
    """
    Return (price_value, price_label) for the most current price in a row.

    Priority:
    1. Look for a header that contains "current reolube price" (Apr 2025 — most recent)
    2. Look for a header that contains "current grease price"  (Oct 2024)
    3. Fall back to the rightmost non-null numeric value > 1 among price columns
       (used for Mail Merge sheets that don't have explicit "Current" columns)
    """
    # Step 1 & 2 – explicit current-price columns
    for keyword, friendly_label in CURRENT_COL_KEYWORDS:
        for h in headers:
            if h is None:
                continue
            if keyword in _normalize(h):
                v = row_orig.get(str(h).strip())
                if v is not None:
                    try:
                        fv = round(float(v), 2)
                        if fv > 1:
                            return fv, friendly_label
                    except (TypeError, ValueError):
                        pass

    # Step 3 – rightmost non-null price column (fallback for Mail Merge sheets)
    last_val   = None
    last_label = None
    for h in headers:
        if not _is_price_col(h):
            continue
        v = row_orig.get(h)
        if v is None:
            continue
        try:
            fv = round(float(v), 2)
            if fv > 1:
                last_val   = fv
                last_label = re.sub(r"\s+", " ", str(h)).strip()
        except (TypeError, ValueError):
            pass

    return last_val, last_label


def _find_col(row_lower: dict, *names: str) -> str:
    """Case-insensitive, whitespace-normalised dict lookup; returns '' if not found."""
    for name in names:
        key = _normalize(name)
        val = row_lower.get(key)
        if val is not None:
            return str(val).strip()
    return ""


def _clean_row(raw_row: tuple, headers: List[str]) -> Optional[Dict]:
    """
    Convert a raw openpyxl row into a clean structured dict:
    customer, product, size, currency, current_price, price_as_of
    """
    row_orig  = {}
    row_lower = {}
    for h, v in zip(headers, raw_row):
        if h is not None and v is not None:
            key_orig  = str(h).strip()
            key_lower = _normalize(key_orig)
            row_orig[key_orig]  = v
            row_lower[key_lower] = v

    customer = _find_col(row_lower, "customer")
    product  = _find_col(row_lower,
                         "product description /code",
                         "product description/code",
                         "product description / code")
    if not customer or not product or customer.lower() in ("blank", "customer"):
        return None

    size     = _find_col(row_lower, "product size")
    currency = _find_col(row_lower, "currency")
    contact  = _find_col(row_lower, "contact email", "contact information")

    current_price, price_label = _current_price(row_orig, headers)

    return {
        "customer":      customer,
        "contact":       contact,
        "product":       product,
        "size":          size,
        "currency":      currency,
        "current_price": current_price,
        "price_as_of":   price_label,
    }


def _parse_excel(content: bytes) -> Dict[str, Any]:
    """
    Parse the price list Excel and return deduplicated clean price records.

    Priority for deduplication (first-write wins — master sheet is most authoritative):
      1. Sept 2024 prices  (master — has explicit Apr 2025 Reolube + Oct 2024 Grease columns)
      2. Mail Merge MOV VSG
      3. Mail merge Grease
      4. Mail merge Reolube  (fallback only — Reolube prices are OLDER than master)
      5. sort               (duplicate of master, skipped if master already covered it)
    """
    import openpyxl
    wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)

    sheet_records: Dict[str, List[Dict]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            continue

        # Find header row (first row with ≥3 non-empty cells)
        header_idx = None
        for i, row in enumerate(all_rows[:10]):
            if sum(1 for c in row if c is not None) >= 3:
                header_idx = i
                break
        if header_idx is None:
            continue

        raw_headers = [str(h).strip() if h is not None else None
                       for h in all_rows[header_idx]]

        records = []
        for raw_row in all_rows[header_idx + 1:]:
            if not any(c is not None for c in raw_row):
                continue
            cleaned = _clean_row(raw_row, raw_headers)
            if cleaned:
                cleaned["_sheet"] = sheet_name
                records.append(cleaned)

        sheet_records[sheet_name] = records
        print(f"[PriceList]   Sheet '{sheet_name}': {len(records)} price records")

    wb.close()

    # Deduplicate: master sheet is processed first; later sheets only fill in gaps.
    # A record already in `seen` with a price is NEVER overwritten.
    seen: Dict[tuple, Dict] = {}
    priority_order = [
        MASTER_SHEET,
        "Mail Merge MOV VSG",
        "Mail merge Grease",
        "Mail merge Reolube",
        "sort",
    ]

    for sheet_name in priority_order:
        for rec in sheet_records.get(sheet_name, []):
            key = (
                rec["customer"].lower(),
                rec["product"].lower().rstrip(),
                rec["size"].lower(),
            )
            if key not in seen:
                seen[key] = rec
            elif seen[key].get("current_price") is None and rec.get("current_price") is not None:
                # Fill in a missing price from a later sheet
                seen[key] = rec

    all_records = list(seen.values())
    print(f"[PriceList] Total unique price records: {len(all_records)}")
    return {
        "records":     all_records,
        "sheet_names": list(sheet_records.keys()),
    }


# ── Public API ────────────────────────────────────────────────────────────────

def load_price_list(force_reload: bool = False) -> Dict[str, Any]:
    """Load (or return cached) price list data."""
    global _cache
    now = time.time()

    if not force_reload and _cache.get("ts") and (now - _cache["ts"]) < CACHE_TTL_SECONDS:
        age = int(now - _cache["ts"])
        print(f"[PriceList] Cache hit ({age}s old, {len(_cache.get('records', []))} records)")
        return _cache

    print("[PriceList] Loading from Google Drive…")
    gdrive_svc = _get_drive_service()
    if not gdrive_svc:
        return {"error": "Google Drive service unavailable", "records": []}

    file_info = _search_file(gdrive_svc)
    if not file_info:
        return {"error": "Price list file not found on Google Drive", "records": []}

    try:
        content = gdrive_svc.download_file(file_info["id"], file_info["name"])
        if content is None or isinstance(content, (dict, list)):
            return {"error": "Failed to download price list file", "records": []}
    except Exception as exc:
        return {"error": f"Download error: {exc}", "records": []}

    parsed = _parse_excel(content)
    if not parsed["records"]:
        return {"error": "No price records found in file", "records": []}

    _cache = {
        "file_name":   file_info["name"],
        "file_id":     file_info["id"],
        "ts":          now,
        "records":     parsed["records"],
        "sheet_names": parsed["sheet_names"],
    }
    return _cache


def search_price(query: str, customer: Optional[str] = None, limit: int = 50) -> Dict[str, Any]:
    """
    Search price records matching the query text and optional customer name.

    Scoring:
    - Each query token found in (customer + product + size) = +1 point
    - Customer name match (partial) = +10 bonus
    - Customer name match (exact)   = +20 bonus
    - Size keyword match (drum/pail/case/keg) = +2 bonus
    """
    data = load_price_list()
    if data.get("error"):
        return {"error": data["error"], "matched_rows": [], "total_matched": 0}

    records = data.get("records", [])
    if not records:
        return {"error": "Price list is empty", "matched_rows": [], "total_matched": 0}

    tokens = [t.lower() for t in re.split(r"[\s,/\-]+", query) if len(t) >= 2]

    def score(rec: dict) -> int:
        searchable = " ".join([
            rec.get("customer", ""),
            rec.get("product",  ""),
            rec.get("size",     ""),
        ]).lower()

        s = sum(1 for t in tokens if t in searchable)
        if s == 0:
            return 0

        if customer:
            cust_lower = customer.lower()
            rec_cust   = rec.get("customer", "").lower()
            if cust_lower == rec_cust:
                s += 20
            elif cust_lower in rec_cust or rec_cust in cust_lower:
                s += 10

        size_lower = rec.get("size", "").lower()
        for size_kw in ["drum", "pail", "case", "keg", "tote", "bag", "kg"]:
            if size_kw in tokens and size_kw in size_lower:
                s += 2

        return s

    scored = [(score(r), r) for r in records]
    scored = [(s, r) for s, r in scored if s > 0]
    scored.sort(key=lambda x: x[0], reverse=True)

    matched = [r for _, r in scored[:limit]]

    # When customer specified, filter to only rows matching that customer (partial match ok)
    if customer and matched:
        cust_lower = customer.lower()
        matched = [
            r for r in matched
            if cust_lower in (r.get("customer") or "").lower()
            or (r.get("customer") or "").lower() in cust_lower
        ]

    # Summarize which currencies exist in the matched rows (for transparency: no conversion)
    currencies_in_data = sorted(set((r.get("currency") or "").strip().upper() for r in matched if (r.get("currency") or "").strip()))
    currencies_in_data = [c for c in currencies_in_data if c]

    return {
        "matched_rows":       matched,
        "total_matched":     len(matched),
        "query":             query,
        "customer":          customer,
        "currencies_available": currencies_in_data,
        "source_file":       data.get("file_name", PRICE_LIST_FILE_NAME),
        "sheet_names":       data.get("sheet_names", []),
    }


def get_all_sheets_summary() -> Dict[str, Any]:
    """Return sheet names, record count, and a sample of records."""
    data = load_price_list()
    if data.get("error"):
        return data
    records = data.get("records", [])
    return {
        "file_name":     data.get("file_name", ""),
        "total_records": len(records),
        "sheet_names":   data.get("sheet_names", []),
        "sample":        records[:5],
        "current_price_columns": {
            "grease_mov_vsg": "Current Grease Price as of Oct '24 increase  (col 21, master sheet)",
            "reolube":        "Current Reolube Price as of Apr '25 increase (col 22, master sheet — MOST RECENT)",
        },
        "note": (
            "Each record = one customer+product+size combination. "
            "current_price is pulled from the explicit 'Current Grease/Reolube' columns "
            "in the master sheet (most authoritative). "
            "Currency is CAD or USD per the customer agreement."
        ),
    }
