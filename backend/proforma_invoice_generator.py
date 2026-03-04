"""
Proforma Invoice Generator - copies the xlsx template and fills data cells.

The template file is NEVER modified. We shutil.copy2 it to the output
location and then open the copy with openpyxl to write only the data cells.
This preserves every aspect of the original template (images, merged cells,
borders, fonts, print settings, headers/footers, column widths, etc.).

Template: backend/templates/proforma_invoice/proforma_invoice_template.xlsx
Output:   Proforma Invoice_CompanyName_PO.xlsx
"""
import os
import shutil
import copy
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


TEMPLATE_DIR = os.path.join(os.path.dirname(__file__), 'templates', 'proforma_invoice')
TEMPLATE_PATH = os.path.join(TEMPLATE_DIR, 'proforma_invoice_template.xlsx')

ITEM_START_ROW = 19
ITEM_END_ROW = 24
MAX_ITEMS = ITEM_END_ROW - ITEM_START_ROW + 1  # 6 (rows 19-24, before Sub Total at row 25)


def _safe_str(val, default=''):
    if val is None:
        return default
    return str(val).strip() or default


def _safe_float(val, default=0.0):
    if val is None:
        return default
    try:
        cleaned = str(val).replace('$', '').replace(',', '').replace('US', '').strip()
        return float(cleaned)
    except (ValueError, TypeError):
        return default


def _safe_int(val, default=0):
    if val is None:
        return default
    try:
        return int(float(str(val).replace(',', '').strip()))
    except (ValueError, TypeError):
        return default


def _parse_date(val):
    """Try to parse a date value into a datetime object."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    s = str(val).strip()
    if not s:
        return None
    for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y-%m-%dT%H:%M:%S',
                '%B %d, %Y', '%b %d, %Y'):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _merge_if_not_merged(ws, cell_range):
    """Merge cells only if they aren't already merged."""
    for existing in ws.merged_cells.ranges:
        if str(existing) == cell_range:
            return
    try:
        ws.merge_cells(cell_range)
    except Exception:
        pass


def generate_proforma_invoice(so_data: dict) -> tuple:
    """
    Copy the proforma invoice template and fill it with sales order data.

    The template is copied as-is first, then only the data cells are written.
    All template formatting, merged cells, formulas (I27 SUM), etc. are preserved.

    Returns:
        (output_path, filename) tuple.
    """
    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(f"Proforma invoice template not found: {TEMPLATE_PATH}")

    customer_name = _safe_str(so_data.get('customer_name'), 'Customer')
    po_number = _safe_str(so_data.get('po_number'))
    so_number = _safe_str(so_data.get('so_number'))
    invoice_date = _parse_date(so_data.get('invoice_date')) or datetime.now()
    date_str = invoice_date.strftime('%Y-%m-%d')

    # --- Build output filename ---
    safe_customer = ''.join(
        c for c in customer_name if c.isalnum() or c in (' ', '-', '_', '.')
    ).strip()
    safe_po = ''.join(
        c for c in po_number if c.isalnum() or c in (' ', '-', '_', '.')
    ).strip()

    if safe_po:
        filename = f'Proforma Invoice_{safe_customer}_{safe_po}.xlsx'
    else:
        filename = f'Proforma Invoice_{safe_customer}_{date_str}.xlsx'

    uploads_dir = os.path.join(os.path.dirname(__file__), 'uploads', 'proforma_invoices')
    os.makedirs(uploads_dir, exist_ok=True)
    output_path = os.path.join(uploads_dir, filename)

    # STEP 1: Copy the template file exactly as-is (binary copy)
    shutil.copy2(TEMPLATE_PATH, output_path)

    # STEP 2: Open the COPY and fill only the data cells
    wb = load_workbook(output_path)
    ws = wb.active

    # Pre-calculate how many extra rows (beyond the 6 template rows) are needed
    # so that row numbers for later sections (trade terms, totals) are correct.
    items = so_data.get('items', [])
    extras = max(0, len(items) - MAX_ITEMS)

    # --- Title row ---
    ws['C4'] = f'Proforma Invoice {date_str}'
    if so_number:
        ws['G4'] = f'SO# {so_number}'

    # --- Buyer / Sold To ---
    ws['C5'] = customer_name
    ws['C7'] = _safe_str(so_data.get('address'))
    ws['C8'] = _safe_str(so_data.get('city_state_zip'))
    ws['C9'] = _safe_str(so_data.get('country'))
    ws['C10'] = _safe_str(so_data.get('phone'))
    ws['C11'] = f' {po_number}' if po_number else ''
    ws['C12'] = _safe_str(so_data.get('email'))

    # --- Shipping / Dates ---
    ship_via = _safe_str(so_data.get('ship_via'))
    if ship_via:
        ws['B15'] = ship_via

    ship_by = _parse_date(so_data.get('ship_by_date'))
    if ship_by:
        ws['E15'] = ship_by

    ws['I15'] = invoice_date

    # --- Trade Terms (row shifts if extra item rows were inserted) ---
    trade_terms = _safe_str(so_data.get('trade_terms'), 'EXW')
    trade_terms_row = 31 + extras
    ws[f'B{trade_terms_row}'] = f'Trade Terms: {trade_terms}'

    # --- Items ---
    # Capture reference formatting from the first template item row
    ref_font_b = copy.copy(ws[f'B{ITEM_START_ROW}'].font)
    ref_font_c = copy.copy(ws[f'C{ITEM_START_ROW}'].font)
    ref_font_g = copy.copy(ws[f'G{ITEM_START_ROW}'].font)
    ref_font_h = copy.copy(ws[f'H{ITEM_START_ROW}'].font)
    ref_font_i = copy.copy(ws[f'I{ITEM_START_ROW}'].font)
    ref_align_h = copy.copy(ws[f'H{ITEM_START_ROW}'].alignment)
    ref_align_i = copy.copy(ws[f'I{ITEM_START_ROW}'].alignment)
    ref_numfmt_h = ws[f'H{ITEM_START_ROW}'].number_format
    ref_numfmt_i = ws[f'I{ITEM_START_ROW}'].number_format

    # If there are more items than the template supports, insert extra rows before
    # the Sub Total row so every line item appears in the output.
    if extras > 0:
        # Insert rows just after the last template item row (before Sub Total)
        ws.insert_rows(ITEM_END_ROW + 1, extras)

    # Recalculate key row numbers after any inserted rows
    actual_end_row = ITEM_END_ROW + extras
    subtotal_row = 25 + extras
    total_row = 27 + extras

    for idx, item in enumerate(items):
        row = ITEM_START_ROW + idx

        _merge_if_not_merged(ws, f'C{row}:F{row}')

        cell_b = ws[f'B{row}']
        cell_b.value = _safe_str(item.get('product_code') or item.get('item_code'))
        cell_b.font = copy.copy(ref_font_b)

        cell_c = ws[f'C{row}']
        cell_c.value = _safe_str(item.get('description'))
        cell_c.font = copy.copy(ref_font_c)

        qty = _safe_int(item.get('quantity'))
        price = _safe_float(item.get('unit_price'))

        cell_g = ws[f'G{row}']
        cell_g.value = qty
        cell_g.font = copy.copy(ref_font_g)

        cell_h = ws[f'H{row}']
        cell_h.value = price
        cell_h.font = copy.copy(ref_font_h)
        cell_h.alignment = copy.copy(ref_align_h)
        cell_h.number_format = ref_numfmt_h or '$#,##0.00'

        cell_i = ws[f'I{row}']
        cell_i.value = f'=H{row}*G{row}'
        cell_i.font = copy.copy(ref_font_i)
        cell_i.alignment = copy.copy(ref_align_i)
        cell_i.number_format = ref_numfmt_i or '$#,##0.00'

    # Subtotal: sum all item rows
    ws[f'I{subtotal_row}'] = f'=SUM(I{ITEM_START_ROW}:I{actual_end_row})'
    # Grand total equals subtotal
    ws[f'I{total_row}'] = f'=I{subtotal_row}'

    wb.save(output_path)
    wb.close()

    return output_path, filename
