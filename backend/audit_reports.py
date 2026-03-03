"""
Audit the inventory month-end data for accuracy.
Check: duplicates, zero-stock items, cost outliers, and compare to Oct 2025 reference.
"""

import json
import csv
import os
import openpyxl
from collections import Counter


def safe_float(v, default=0.0):
    if v is None:
        return default
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s or s == 'None' or s == 'False':
        return default
    try:
        return float(s.replace(',', '').replace('$', '').strip())
    except (TypeError, ValueError):
        return default


def audit_jan():
    print("=" * 70)
    print("  AUDIT: JANUARY 2026 DATA (CustomAlert5.json from Jan 27)")
    print("=" * 70)

    path = r"G:\Shared drives\IT_Automation\MiSys\Misys Extracted Data\API Extractions\2026-01-27\CustomAlert5.json"
    with open(path, 'r', encoding='utf-8') as f:
        items = json.load(f)

    print(f"\nTotal raw records: {len(items)}")

    # Check for duplicates
    item_nos = [str(i.get('Item No.') or '').strip() for i in items if isinstance(i, dict)]
    dupes = {k: v for k, v in Counter(item_nos).items() if v > 1 and k}
    print(f"Duplicate Item No. entries: {len(dupes)}")
    if dupes:
        for k, v in sorted(dupes.items(), key=lambda x: -x[1])[:10]:
            print(f"  '{k}' appears {v} times")

    # Break down by stock status
    zero_stock = 0
    negative_stock = 0
    positive_stock = 0
    no_cost = 0
    zero_value = 0
    total_value = 0
    total_qty = 0
    positive_value_items = []

    for item in items:
        if not isinstance(item, dict):
            continue
        item_no = str(item.get('Item No.') or '').strip()
        stock = safe_float(item.get('Stock'))
        unit_cost = safe_float(item.get('Recent Cost'))
        if not unit_cost:
            unit_cost = safe_float(item.get('Unit Cost'))
        if not unit_cost:
            unit_cost = safe_float(item.get('Average Cost'))
        if not unit_cost:
            unit_cost = safe_float(item.get('Standard Cost'))
        ext = stock * unit_cost

        if stock == 0:
            zero_stock += 1
        elif stock < 0:
            negative_stock += 1
        else:
            positive_stock += 1

        if unit_cost == 0:
            no_cost += 1
        if ext == 0:
            zero_value += 1

        total_value += ext
        total_qty += stock

        if ext != 0:
            positive_value_items.append({
                'item_no': item_no,
                'desc': (item.get('Description') or '')[:40],
                'stock': stock,
                'cost': unit_cost,
                'value': ext,
            })

    print(f"\nStock breakdown:")
    print(f"  Items with POSITIVE stock: {positive_stock}")
    print(f"  Items with ZERO stock:     {zero_stock}")
    print(f"  Items with NEGATIVE stock: {negative_stock}")
    print(f"  Items with NO unit cost:   {no_cost}")
    print(f"  Items with $0 value:       {zero_value}")

    print(f"\nIf we ONLY count items with stock > 0:")
    pos_items = [i for i in positive_value_items if i['stock'] > 0]
    pos_total = sum(i['value'] for i in pos_items)
    print(f"  Items: {len(pos_items)}")
    print(f"  Total Value: ${pos_total:,.2f}")

    print(f"\nIf we count ALL items (including zero stock):")
    print(f"  Items: {len(items)}")
    print(f"  Total Value: ${total_value:,.2f}")

    print(f"\nItems with value != 0: {len(positive_value_items)}")
    print(f"Their total value: ${sum(i['value'] for i in positive_value_items):,.2f}")

    # Top 15 by value
    positive_value_items.sort(key=lambda x: -abs(x['value']))
    print(f"\nTOP 15 items by $$ value:")
    print(f"  {'Item No.':<25} {'Description':<40} {'On Hand':>12} {'Unit Cost':>12} {'Value':>14}")
    print(f"  {'-'*25} {'-'*40} {'-'*12} {'-'*12} {'-'*14}")
    for i in positive_value_items[:15]:
        print(f"  {i['item_no']:<25} {i['desc']:<40} {i['stock']:>12,.2f} ${i['cost']:>11,.4f} ${i['value']:>13,.2f}")

    # Bottom 15 (smallest non-zero)
    nonzero = [i for i in positive_value_items if i['value'] > 0]
    nonzero.sort(key=lambda x: x['value'])
    print(f"\n  BOTTOM 5 items (smallest positive value):")
    for i in nonzero[:5]:
        print(f"  {i['item_no']:<25} {i['desc']:<40} {i['stock']:>12,.2f} ${i['cost']:>11,.4f} ${i['value']:>13,.2f}")

    # Cost outliers (very high cost per unit)
    high_cost = sorted([i for i in positive_value_items if i['cost'] > 100], key=lambda x: -x['cost'])
    print(f"\n  HIGH UNIT COST items (>$100/unit): {len(high_cost)}")
    for i in high_cost[:10]:
        print(f"  {i['item_no']:<25} {i['desc']:<40} cost=${i['cost']:>10,.2f}  value=${i['value']:>12,.2f}")

    return total_value, len(items), positive_stock


def audit_feb():
    print("\n\n" + "=" * 70)
    print("  AUDIT: FEBRUARY 2026 DATA (MIITEM.CSV from Feb 23)")
    print("=" * 70)

    feb_base = r"G:\Shared drives\IT_Automation\MiSys\Misys Extracted Data\Full Company Data From Misys\Feb 23, 2026"

    items = []
    with open(os.path.join(feb_base, "MIITEM.CSV"), 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            items.append(row)

    print(f"\nTotal raw records: {len(items)}")

    item_nos = [str(r.get('itemId') or '').strip().strip('"') for r in items]
    dupes = {k: v for k, v in Counter(item_nos).items() if v > 1 and k}
    print(f"Duplicate itemId entries: {len(dupes)}")
    if dupes:
        for k, v in sorted(dupes.items(), key=lambda x: -x[1])[:10]:
            print(f"  '{k}' appears {v} times")

    zero_stock = 0
    negative_stock = 0
    positive_stock = 0
    no_cost = 0
    total_value = 0
    total_qty = 0
    positive_value_items = []

    for row in items:
        item_no = str(row.get('itemId') or '').strip().strip('"')
        stock = safe_float(row.get('totQStk'))
        unit_cost = safe_float(row.get('cLast'))
        if not unit_cost:
            unit_cost = safe_float(row.get('itemCost'))
        if not unit_cost:
            unit_cost = safe_float(row.get('cAvg'))
        if not unit_cost:
            unit_cost = safe_float(row.get('cStd'))
        ext = stock * unit_cost
        desc = str(row.get('descr') or '')[:40].strip().strip('"')

        if stock == 0:
            zero_stock += 1
        elif stock < 0:
            negative_stock += 1
        else:
            positive_stock += 1

        if unit_cost == 0:
            no_cost += 1

        total_value += ext
        total_qty += stock

        if ext != 0:
            positive_value_items.append({
                'item_no': item_no, 'desc': desc,
                'stock': stock, 'cost': unit_cost, 'value': ext,
            })

    print(f"\nStock breakdown:")
    print(f"  Items with POSITIVE stock: {positive_stock}")
    print(f"  Items with ZERO stock:     {zero_stock}")
    print(f"  Items with NEGATIVE stock: {negative_stock}")
    print(f"  Items with NO unit cost:   {no_cost}")

    print(f"\nIf we ONLY count items with stock > 0:")
    pos_items = [i for i in positive_value_items if i['stock'] > 0]
    pos_total = sum(i['value'] for i in pos_items)
    print(f"  Items: {len(pos_items)}")
    print(f"  Total Value: ${pos_total:,.2f}")

    print(f"\nAll items total value: ${total_value:,.2f}")
    print(f"Items with value != 0: {len(positive_value_items)}")

    positive_value_items.sort(key=lambda x: -abs(x['value']))
    print(f"\nTOP 15 items by $$ value:")
    print(f"  {'Item No.':<25} {'Description':<40} {'On Hand':>12} {'Unit Cost':>12} {'Value':>14}")
    print(f"  {'-'*25} {'-'*40} {'-'*12} {'-'*12} {'-'*14}")
    for i in positive_value_items[:15]:
        print(f"  {i['item_no']:<25} {i['desc']:<40} {i['stock']:>12,.2f} ${i['cost']:>11,.4f} ${i['value']:>13,.2f}")

    high_cost = sorted([i for i in positive_value_items if i['cost'] > 100], key=lambda x: -x['cost'])
    print(f"\n  HIGH UNIT COST items (>$100/unit): {len(high_cost)}")
    for i in high_cost[:10]:
        print(f"  {i['item_no']:<25} {i['desc']:<40} cost=${i['cost']:>10,.2f}  value=${i['value']:>12,.2f}")

    return total_value, len(items), positive_stock


def audit_oct_reference():
    print("\n\n" + "=" * 70)
    print("  REFERENCE: OCTOBER 2025 FILE")
    print("=" * 70)

    path = r"C:\Users\Haron\Downloads\Inventory Month End for October 2025.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Stock Transfers']

    print(f"\nSheet: 'Stock Transfers'")
    print(f"Headers: {[ws.cell(1, c).value for c in range(1, ws.max_column+1)]}")
    print(f"Data rows: {ws.max_row - 1}")

    unique_items = set()
    total_cost_out = 0
    total_cost_in = 0
    categories = Counter()

    for r in range(2, ws.max_row + 1):
        cat = ws.cell(r, 2).value or ''
        item = ws.cell(r, 4).value or ''
        cost_out = float(ws.cell(r, 6).value or 0)
        cost_in = float(ws.cell(r, 7).value or 0)
        categories[cat] += 1
        if item and item != 'TOTAL':
            unique_items.add(item)
        total_cost_out += cost_out
        total_cost_in += cost_in

    print(f"Unique items in Oct file: {len(unique_items)}")
    print(f"Categories: {dict(categories)}")
    print(f"Total Cost Out: ${total_cost_out:,.2f}")
    print(f"Total Cost In:  ${total_cost_in:,.2f}")
    print(f"\nNOTE: This file is STOCK TRANSFERS, not inventory on hand.")
    print(f"      It shows movement of goods, not ending balances.")


if __name__ == "__main__":
    jan_val, jan_cnt, jan_pos = audit_jan()
    feb_val, feb_cnt, feb_pos = audit_feb()
    audit_oct_reference()

    print("\n\n" + "=" * 70)
    print("  RECOMMENDATION")
    print("=" * 70)
    print(f"\n  January: {jan_cnt} total items, but only {jan_pos} have stock on hand")
    print(f"  February: {feb_cnt} total items, but only {feb_pos} have stock on hand")
    print(f"\n  The reports currently include ALL items (even zero-stock).")
    print(f"  Should we filter to ONLY items with stock on hand > 0?")
    print(f"  That would significantly reduce the item count.")
