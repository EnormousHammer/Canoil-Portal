"""Verify generated reports match the Oct 2025 reference format."""
import openpyxl

ref_path = r"C:\Users\Haron\Downloads\Inventory Month End for October 2025.xlsx"
jan_path = r"G:\Shared drives\IT_Automation\Canoil Apps\Canoil Helper\canoil-portal\reports\Inventory Month End for January 2026.xlsx"

print("=" * 65)
print("  FORMAT COMPARISON: Reference vs Generated")
print("=" * 65)

# Reference
wb_ref = openpyxl.load_workbook(ref_path)
ws_ref = wb_ref.active
print(f"\n--- REFERENCE (Oct 2025) ---")
print(f"  Sheet name: '{ws_ref.title}'")
print(f"  Headers: {[c.value for c in ws_ref[1]]}")
print(f"  Total data rows: {ws_ref.max_row - 1}")
print(f"  Column widths: {[round(ws_ref.column_dimensions[chr(65+i)].width or 0, 1) for i in range(7)]}")

# First 3 data rows
print("  First 3 data rows:")
for row in ws_ref.iter_rows(min_row=2, max_row=4, values_only=False):
    vals = []
    for c in row:
        v = c.value
        if hasattr(v, 'strftime'):
            v = v.strftime('%Y-%m-%d')
        elif isinstance(v, float):
            v = f"{v:.5f}"
        vals.append(str(v) if v else '')
    print(f"    {vals}")

# Last row (TOTAL)
print("  Last row (TOTAL):")
last = list(ws_ref.iter_rows(min_row=ws_ref.max_row, max_row=ws_ref.max_row, values_only=True))[0]
print(f"    {list(last)}")

# Categories in reference
cats = set()
for row in ws_ref.iter_rows(min_row=2, max_row=ws_ref.max_row-1, values_only=True):
    if row[1]: cats.add(row[1])
print(f"  Categories found: {sorted(cats)}")

# Generated
wb_jan = openpyxl.load_workbook(jan_path)
ws_jan = wb_jan.active
print(f"\n--- GENERATED (Jan 2026) ---")
print(f"  Sheet name: '{ws_jan.title}'")
print(f"  Headers: {[c.value for c in ws_jan[1]]}")
print(f"  Total data rows: {ws_jan.max_row - 1}")

# First 3 data rows
print("  First 3 data rows:")
for row in ws_jan.iter_rows(min_row=2, max_row=4, values_only=False):
    vals = []
    for c in row:
        v = c.value
        if hasattr(v, 'strftime'):
            v = v.strftime('%Y-%m-%d')
        elif isinstance(v, float):
            v = f"{v:.5f}"
        vals.append(str(v) if v else '')
    print(f"    {vals}")

# Last row (TOTAL)
print("  Last row (TOTAL):")
last = list(ws_jan.iter_rows(min_row=ws_jan.max_row, max_row=ws_jan.max_row, values_only=True))[0]
print(f"    {list(last)}")

# Categories
cats_jan = set()
for row in ws_jan.iter_rows(min_row=2, max_row=ws_jan.max_row-1, values_only=True):
    if row[1]: cats_jan.add(row[1])
print(f"  Categories found: {sorted(cats_jan)}")

print("\n--- FORMAT MATCH CHECK ---")
ref_hdrs = [c.value for c in ws_ref[1]]
jan_hdrs = [c.value for c in ws_jan[1]]
print(f"  Headers match: {'YES' if ref_hdrs == jan_hdrs else 'NO'}")
print(f"    Ref: {ref_hdrs}")
print(f"    Gen: {jan_hdrs}")
print(f"  Sheet name match: {'YES' if ws_ref.title == ws_jan.title else 'NO'}")
print(f"  Categories match: {'YES' if cats == cats_jan else 'NO - ref has ' + str(sorted(cats)) + ', gen has ' + str(sorted(cats_jan))}")

# Check number format on Cost columns
ref_fmt6 = ws_ref.cell(row=2, column=6).number_format
gen_fmt6 = ws_jan.cell(row=2, column=6).number_format
print(f"  Cost Out format: ref='{ref_fmt6}' gen='{gen_fmt6}'")

# Bold headers?
ref_bold = ws_ref.cell(row=1, column=1).font.bold
gen_bold = ws_jan.cell(row=1, column=1).font.bold
print(f"  Bold headers: ref={ref_bold}, gen={gen_bold}")

# TOTAL row bold?
ref_total_bold = ws_ref.cell(row=ws_ref.max_row, column=5).font.bold
gen_total_bold = ws_jan.cell(row=ws_jan.max_row, column=5).font.bold
print(f"  Bold TOTAL: ref={ref_total_bold}, gen={gen_total_bold}")

print("\n" + "=" * 65)
