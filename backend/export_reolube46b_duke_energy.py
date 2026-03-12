#!/usr/bin/env python3
"""Export REOLUBE46B to Duke Energy orders to Excel. Cost matched to order date from MiSys MIICST (cost history) + MIPOD (PO receipts). FX = Bank of Canada historical USD/CAD."""
import csv
import json
import os
import re
import urllib.request
from datetime import datetime
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

SAGE_BASE = os.environ.get(
    "SAGE_GDRIVE_BASE",
    r"G:\Shared drives\IT_Automation\MiSys\Misys Extracted Data\Full Company Data From SAGE"
)
MISYS_BASE = os.environ.get(
    "FULL_COMPANY_DATA_BASE",
    r"G:\Shared drives\IT_Automation\MiSys\Misys Extracted Data\Full Company Data From Misys"
)
USD_CAD_RATE = float(os.environ.get("USD_CAD_RATE", "1.36"))  # 1 USD = X CAD

def _parse_date(val):
    """Parse date from various formats: YYYY-MM-DD, YYYYMMDD, datetime string."""
    if not val:
        return None
    s = str(val).strip()[:20]
    # YYYY-MM-DD
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        try:
            return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            pass
    # YYYYMMDD
    m = re.match(r"(\d{4})(\d{2})(\d{2})", s)
    if m:
        try:
            return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            pass
    return None

def _load_csv(folder_path, table_name):
    for name in [f"{table_name}.CSV", f"{table_name}.csv"]:
        fp = os.path.join(folder_path, name)
        if os.path.isfile(fp):
            rows = []
            with open(fp, encoding="utf-8-sig", newline="", errors="replace") as f:
                for r in csv.DictReader(f):
                    rows.append(r)
            return rows
    return []

def _safe_float(val, default=0.0):
    if val is None or val == "":
        return default
    try:
        s = str(val).replace("$", "").replace(",", "").strip()
        return float(s) if s else default
    except Exception:
        return default

def _safe_str(val):
    return str(val or "").strip()

def _load_boc_usdcad_rates(start_date, end_date):
    """Fetch Bank of Canada USD/CAD FX rates. Returns (rates_dict, sorted_dates)."""
    if start_date is None or end_date is None:
        return {}, []
    s = start_date.strftime("%Y-%m-%d")
    e = end_date.strftime("%Y-%m-%d")
    url = f"https://www.bankofcanada.ca/valet/observations/FXUSDCAD/json?start_date={s}&end_date={e}"
    rates = {}
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Canoil-Portal/1.0"})
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode())
        for o in data.get("observations", []):
            d = o.get("d")
            v = o.get("FXUSDCAD", {}).get("v")
            if d and v is not None:
                try:
                    rates[d] = float(v)
                except (TypeError, ValueError):
                    pass
    except Exception:
        pass
    return rates, sorted(rates.keys())

def _get_fx_rate(order_date, boc_rates, boc_sorted_dates, sage_rate, fallback_rate):
    """Get USD->CAD rate: BOC > Sage dExchRate > fallback."""
    if order_date is None:
        return fallback_rate
    dt_str = order_date.strftime("%Y-%m-%d")
    if dt_str in boc_rates:
        return boc_rates[dt_str]
    for d in reversed(boc_sorted_dates):
        if d <= dt_str:
            return boc_rates[d]
    if sage_rate and float(sage_rate) > 0:
        return float(sage_rate)
    return fallback_rate

def _load_misys_cost_history():
    """
    Load cost history with CURRENCY. PO cost can be USD (LANXESS USD) or CAD.
    Returns: (cost_history, current_cost_usd, current_cost_cad)
    - cost_history: itemId -> [(date, cost_usd, cost_cad)] sorted by date
    - current_cost_usd/cad: fallback from MIITEM
    """
    if not os.path.isdir(MISYS_BASE):
        return {}, {}, {}
    subfolders = [f for f in os.listdir(MISYS_BASE)
                  if os.path.isdir(os.path.join(MISYS_BASE, f)) and not f.startswith("_")]
    if not subfolders:
        return {}, {}, {}
    subfolders.sort(key=lambda f: os.path.getmtime(os.path.join(MISYS_BASE, f)), reverse=True)
    folder = os.path.join(MISYS_BASE, subfolders[0])

    # MIPOH: pohId -> (srcCur, rate) for PO currency
    mipoh = _load_csv(folder, "MIPOH")
    po_currency = {}
    for r in mipoh:
        poh = _safe_str(r.get("pohId") or r.get("PO No."))
        src = (_safe_str(r.get("srcCur") or "") or "CAD").upper()
        rate = _safe_float(r.get("rate") or 1)
        if poh:
            po_currency[poh] = (src, rate if rate > 0 else USD_CAD_RATE)

    # cost_history: itemId -> [(date, cost_usd, cost_cad)]
    cost_history = {}
    seen = {}  # (key, dt) -> (cost_usd, cost_cad, is_po)

    # MIICST: use "cost" (product cost per drum) only — NOT cLand (landed cost with freight/duties)
    miicst = _load_csv(folder, "MIICST")
    for r in miicst:
        item_id = _safe_str(r.get("itemId") or r.get("Item No."))
        if not item_id:
            continue
        cost_cad = _safe_float(r.get("cost"))  # product cost only; cLand includes freight/duties
        if cost_cad <= 0:
            continue
        dt = _parse_date(r.get("transDt") or r.get("transDate"))
        if not dt:
            continue
        key = item_id.upper()
        cost_usd = cost_cad / USD_CAD_RATE
        tt = int(r.get("tranType") or r.get("type") or 0)
        is_po = (tt == 0)
        dkey = (key, dt)
        if dkey not in seen or (is_po and not seen[dkey][2]):
            seen[dkey] = (cost_usd, cost_cad, is_po)

    # MIPOD: PO price - currency from MIPOH srcCur
    # MIPOD may store unit cost OR extended (line total). If cost/qty is a reasonable drum price, use it.
    mipod = _load_csv(folder, "MIPOD")
    for r in mipod:
        item_id = _safe_str(r.get("itemId") or r.get("Item No."))
        if not item_id:
            continue
        cost = _safe_float(r.get("cost") or r.get("price"))
        if cost <= 0:
            continue
        qty = _safe_float(r.get("ordered") or r.get("Ordered") or r.get("received") or r.get("Received") or 0)
        if qty > 0:
            unit_cost = cost / qty
            if 2000 <= unit_cost <= 10000:  # reasonable 230kg drum range
                cost = unit_cost
        dt = _parse_date(r.get("lastRecvDt") or r.get("lastRecvDate") or r.get("realDueDt") or r.get("initDueDt"))
        if not dt:
            continue
        poh = _safe_str(r.get("pohId") or r.get("PO No."))
        src_cur, rate = po_currency.get(poh, ("CAD", USD_CAD_RATE))
        if src_cur == "USD":
            cost_usd, cost_cad = cost, cost * rate
        else:
            cost_usd, cost_cad = cost / USD_CAD_RATE, cost
        key = item_id.upper()
        dkey = (key, dt)
        # MIPOD has actual PO currency - prefer over MIICST for same date
        seen[dkey] = (cost_usd, cost_cad, True)

    for (key, dt), (cu, cc, _) in seen.items():
        if key not in cost_history:
            cost_history[key] = []
        cost_history[key].append((dt, cu, cc))
    for key in cost_history:
        cost_history[key].sort(key=lambda x: x[0])

    # MIITEM fallback (assume CAD for cLast/cStd)
    miitem = _load_csv(folder, "MIITEM")
    current_usd, current_cad = {}, {}
    for r in miitem:
        item_id = _safe_str(r.get("itemId") or r.get("Item No."))
        if not item_id:
            continue
        c = _safe_float(r.get("cLast") or r.get("cStd"))
        if c > 0:
            k = item_id.upper()
            current_cad[k] = c
            current_usd[k] = c / USD_CAD_RATE

    return cost_history, current_usd, current_cad

def _cost_as_of_date(cost_history, current_usd, current_cad, item_code, order_date_str):
    """Return (cost_usd, cost_cad) for item as of order_date."""
    key = (item_code or "").upper()
    if not key:
        return 0, 0
    order_dt = _parse_date(order_date_str)
    if cost_history.get(key):
        lst = cost_history[key]
        if not order_dt:
            cu, cc = lst[-1][1], lst[-1][2]
            return cu, cc
        for dt, cu, cc in reversed(lst):
            if dt <= order_dt:
                return cu, cc
    return current_usd.get(key, 0) or 0, current_cad.get(key, 0) or 0

def main():
    if not os.path.isdir(SAGE_BASE):
        print(f"ERROR: Sage folder not found: {SAGE_BASE}")
        return None
    subfolders = [f for f in os.listdir(SAGE_BASE)
                 if os.path.isdir(os.path.join(SAGE_BASE, f)) and not f.startswith("_")]
    if not subfolders:
        print(f"ERROR: No subfolders")
        return None
    subfolders.sort(key=lambda f: os.path.getmtime(os.path.join(SAGE_BASE, f)), reverse=True)
    folder = os.path.join(SAGE_BASE, subfolders[0])

    tsalordr = _load_csv(folder, "tsalordr")
    tsoline = _load_csv(folder, "tsoline")
    tinvent = _load_csv(folder, "tinvent")
    tcustomr = _load_csv(folder, "tcustomr")

    if not tsalordr or not tsoline or not tinvent:
        print("ERROR: Missing data")
        return None

    cost_history, current_usd, current_cad = _load_misys_cost_history()

    cust_search = "Duke Energy"
    item_search = "REOLUBE46B"
    cust_lower = cust_search.lower()
    item_variants = ["reolube46b", "reol46b", "reol46bdrm"]

    cust_name_map = {int(r.get("lId", 0) or 0): _safe_str(r.get("sName")) for r in tcustomr}
    cust_ids = [cid for cid, name in cust_name_map.items() if cust_lower in (name or "").lower()]
    if not cust_ids:
        print("No customers found")
        return None

    so_ids = set()
    so_id_to_header = {}
    CURRENCY_MAP = {1: "CAD", 2: "USD"}
    for r in tsalordr:
        cid = int(r.get("lCusId", 0) or 0)
        sid = int(r.get("lId", 0) or 0)
        if cid in cust_ids:
            so_ids.add(sid)
            cur_id = int(r.get("lCurrncyId", 1) or 1)
            dt_str = (_safe_str(r.get("dtSODate")) or "")[:10]
            so_id_to_header[sid] = {
                "sSONum": _safe_str(r.get("sSONum")),
                "dtSODate": _safe_str(r.get("dtSODate")),
                "lCusId": cid,
                "currency": CURRENCY_MAP.get(cur_id, "CAD"),
                "dExchRate": _safe_float(r.get("dExchRate") or 0),
            }

    order_dates = [_parse_date((h.get("dtSODate") or "")[:10]) for h in so_id_to_header.values()]
    valid_dates = [d for d in order_dates if d is not None]
    boc_rates, boc_sorted_dates = _load_boc_usdcad_rates(min(valid_dates), max(valid_dates)) if valid_dates else ({}, [])

    inv_map = {}
    for r in tinvent:
        iid = int(r.get("lId", 0) or 0)
        inv_map[iid] = {"sPartCode": _safe_str(r.get("sPartCode")), "sName": _safe_str(r.get("sName"))}

    records = []
    for row in tsoline:
        so_id = int(row.get("lSOId", 0) or 0)
        if so_id not in so_ids:
            continue
        iid = int(row.get("lInventId", 0) or 0)
        info = inv_map.get(iid, {"sPartCode": "", "sName": ""})
        part = info["sPartCode"].lower()
        name = (info["sName"] or "").lower()
        if not any(v in part or v in name for v in item_variants):
            continue
        hdr = so_id_to_header.get(so_id, {})
        qty = _safe_float(row.get("dQuantity") or row.get("dOrdered") or row.get("dQty") or 0)
        amt = _safe_float(row.get("dAmount") or 0)
        price = _safe_float(row.get("dPrice") or 0)
        if amt == 0 and price > 0 and qty > 0:
            amt = price * qty
        if qty == 0 and amt > 0 and price > 0:
            qty = amt / price
        unit_price = price or (amt / qty if qty > 0 else 0)
        item_code = info["sPartCode"]
        order_date_str = (hdr.get("dtSODate") or "")[:10]
        order_dt = _parse_date(order_date_str)
        cost_usd, _ = _cost_as_of_date(cost_history, current_usd, current_cad, item_code, order_date_str)
        rate = _get_fx_rate(order_dt, boc_rates, boc_sorted_dates, hdr.get("dExchRate"), USD_CAD_RATE)
        cost_cad = cost_usd * rate if cost_usd > 0 else 0
        cur = hdr.get("currency", "CAD")
        rev_usd = amt if cur == "USD" else amt / rate
        rev_cad = amt if cur == "CAD" else amt * rate
        tc_usd = round(cost_usd * qty, 2) if cost_usd > 0 else 0
        tc_cad = round(cost_cad * qty, 2) if cost_cad > 0 else 0
        # Margin = (R - C) / R. With cost_cad = cost_usd * rate, margin is identical: (R*rate - C*rate)/(R*rate) = (R-C)/R
        margin_pct = round((rev_usd - tc_usd) / rev_usd * 100, 1) if rev_usd > 0 and tc_usd > 0 else None
        margin_usd = margin_cad = margin_pct
        profit_usd = round(rev_usd - tc_usd, 2) if rev_usd > 0 and tc_usd > 0 else None
        profit_cad = round(rev_cad - tc_cad, 2) if rev_cad > 0 and tc_cad > 0 else None
        records.append({
            "so_number": hdr.get("sSONum", ""),
            "customer_name": cust_name_map.get(hdr.get("lCusId", 0), ""),
            "item_code": item_code,
            "quantity": qty,
            "unit_price": unit_price,
            "line_total": amt,
            "currency": cur,
            "order_date": (hdr.get("dtSODate") or "")[:10],
            "rev_usd": rev_usd,
            "rev_cad": rev_cad,
            "cost_usd": cost_usd if cost_usd > 0 else None,
            "cost_cad": cost_cad if cost_cad > 0 else None,
            "total_cost_usd": tc_usd if tc_usd > 0 else None,
            "total_cost_cad": tc_cad if tc_cad > 0 else None,
            "profit_usd": profit_usd,
            "profit_cad": profit_cad,
            "margin_usd": margin_usd,
            "margin_cad": margin_cad,
        })

    records.sort(key=lambda x: x["order_date"], reverse=True)

    wb = Workbook()
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    def _write_sheet(ws, title, curr, rev_col, cost_col, tc_col, profit_col, margin_pct_col):
        ws.title = title
        ws.merge_cells("A1:K1")
        ws["A1"] = f"REOLUBE46B Duke Energy — ALL IN {curr}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws["A2"] = f"Report generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | Cost = product cost per drum (230kg), excluding freight/duties | FX = Bank of Canada historical USD/CAD (fallback: Sage dExchRate)"
        ws["A2"].font = Font(italic=True, size=9)
        ws.merge_cells("A2:K2")
        hdrs = ["SO #", "Customer", "Item", "Qty", "Unit Price", "Line Total", "Cost (per drum)", "Total Cost", "Profit", "Margin %", "Order Date"]
        for col, h in enumerate(hdrs, 1):
            c = ws.cell(row=4, column=col, value=h)
            c.font = header_font_white
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = thin_border
        for row_idx, r in enumerate(records, 5):
            ws.cell(row=row_idx, column=1, value=r["so_number"]).border = thin_border
            ws.cell(row=row_idx, column=2, value=r["customer_name"]).border = thin_border
            ws.cell(row=row_idx, column=3, value=r["item_code"]).border = thin_border
            ws.cell(row=row_idx, column=4, value=r["quantity"]).border = thin_border
            q = r["quantity"] or 1
            up = (r[rev_col] / q) if q > 0 else 0
            ws.cell(row=row_idx, column=5, value=up).number_format = '"$"#,##0.00'
            ws.cell(row=row_idx, column=5).border = thin_border
            ws.cell(row=row_idx, column=6, value=r[rev_col]).number_format = '"$"#,##0.00'
            ws.cell(row=row_idx, column=6).border = thin_border
            cv = r.get(cost_col)
            ws.cell(row=row_idx, column=7, value=cv).number_format = '"$"#,##0.00'
            ws.cell(row=row_idx, column=7).border = thin_border
            tcv = r.get(tc_col)
            ws.cell(row=row_idx, column=8, value=tcv).number_format = '"$"#,##0.00'
            ws.cell(row=row_idx, column=8).border = thin_border
            pv = r.get(profit_col)
            ws.cell(row=row_idx, column=9, value=pv).number_format = '"$"#,##0.00'
            ws.cell(row=row_idx, column=9).border = thin_border
            mpv = r.get(margin_pct_col)
            ws.cell(row=row_idx, column=10, value=mpv).number_format = '0.0"%"'
            ws.cell(row=row_idx, column=10).border = thin_border
            ws.cell(row=row_idx, column=11, value=r["order_date"]).border = thin_border
        sr = len(records) + 6
        tot_rev = sum(r[rev_col] for r in records)
        tot_cost = sum(r.get(tc_col) or 0 for r in records)
        tot_profit = sum(r.get(profit_col) or 0 for r in records)
        tot_margin_pct = round((tot_profit / tot_rev * 100), 1) if tot_rev > 0 else None
        ws.cell(row=sr, column=1, value="SUMMARY").font = Font(bold=True)
        ws.cell(row=sr + 1, column=1, value=f"Total Orders: {len(records)}")
        ws.cell(row=sr + 2, column=1, value=f"Total Quantity: {sum(r['quantity'] for r in records)}")
        ws.cell(row=sr + 3, column=1, value=f"Total Revenue: ${tot_rev:,.2f} ({curr})").font = Font(bold=True)
        ws.cell(row=sr + 4, column=1, value=f"Total Cost: ${tot_cost:,.2f} ({curr})").font = Font(bold=True)
        ws.cell(row=sr + 5, column=1, value=f"Total Profit: ${tot_profit:,.2f} ({curr})").font = Font(bold=True)
        ws.cell(row=sr + 6, column=1, value=f"Total Margin: {tot_margin_pct}%" if tot_margin_pct is not None else "Total Margin: —").font = Font(bold=True)
        return sr

    ws_usd = wb.active
    _write_sheet(ws_usd, "USD", "USD", "rev_usd", "cost_usd", "total_cost_usd", "profit_usd", "margin_usd")
    ws_cad = wb.create_sheet("CAD")
    summary_row = _write_sheet(ws_cad, "CAD", "CAD", "rev_cad", "cost_cad", "total_cost_cad", "profit_cad", "margin_cad")

    # By-year aggregates
    by_year = {}
    for r in records:
        yr = (r["order_date"] or "")[:4]
        if not yr:
            continue
        if yr not in by_year:
            by_year[yr] = {"rev_usd": 0, "cost_usd": 0, "rev_cad": 0, "cost_cad": 0, "qty": 0}
        by_year[yr]["rev_usd"] += r["rev_usd"]
        by_year[yr]["cost_usd"] += r.get("total_cost_usd") or 0
        by_year[yr]["rev_cad"] += r["rev_cad"]
        by_year[yr]["cost_cad"] += r.get("total_cost_cad") or 0
        by_year[yr]["qty"] += r["quantity"] or 0

    # Cost per unit & Sell price per unit BY YEAR (for each currency)
    def _add_by_year_table(ws, start_row, title):
        row = start_row
        ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=11)
        row += 1
        ws.cell(row=row, column=1, value="Year")
        ws.cell(row=row, column=2, value="Cost/unit (USD)")
        ws.cell(row=row, column=3, value="Sell/unit (USD)")
        ws.cell(row=row, column=4, value="Cost/unit (CAD)")
        ws.cell(row=row, column=5, value="Sell/unit (CAD)")
        ws.cell(row=row, column=6, value="Margin %")
        row += 1
        for yr in sorted(by_year.keys()):
            d = by_year[yr]
            q = d["qty"] or 1
            cost_usd = d["cost_usd"] / q
            sell_usd = d["rev_usd"] / q
            cost_cad = d["cost_cad"] / q
            sell_cad = d["rev_cad"] / q
            profit_usd = d["rev_usd"] - d["cost_usd"]
            margin_pct = round((profit_usd / d["rev_usd"] * 100), 1) if d["rev_usd"] > 0 else None
            ws.cell(row=row, column=1, value=yr)
            ws.cell(row=row, column=2, value=round(cost_usd, 2)).number_format = '"$"#,##0.00'
            ws.cell(row=row, column=3, value=round(sell_usd, 2)).number_format = '"$"#,##0.00'
            ws.cell(row=row, column=4, value=round(cost_cad, 2)).number_format = '"$"#,##0.00'
            ws.cell(row=row, column=5, value=round(sell_cad, 2)).number_format = '"$"#,##0.00'
            mc = ws.cell(row=row, column=6, value=margin_pct if margin_pct is not None else "—")
            if margin_pct is not None:
                mc.number_format = '0.0"%"'
            row += 1
        return row

    by_year_start = summary_row + 8
    by_year_end = _add_by_year_table(ws_usd, by_year_start, "BY YEAR: Cost per unit (what we pay) & Sell price per unit (what we charge)")
    _add_by_year_table(ws_cad, by_year_start, "BY YEAR: Cost per unit (what we pay) & Sell price per unit (what we charge)")

    chart_start = by_year_end + 2
    ws_cad.cell(row=chart_start, column=1, value="Chart Data (totals by year)").font = Font(bold=True)
    chart_start += 1
    ws_cad.cell(row=chart_start, column=1, value="Year")
    ws_cad.cell(row=chart_start, column=2, value="Revenue (USD)")
    ws_cad.cell(row=chart_start, column=3, value="Cost (USD)")
    ws_cad.cell(row=chart_start, column=4, value="Revenue (CAD)")
    ws_cad.cell(row=chart_start, column=5, value="Cost (CAD)")
    chart_start += 1
    for yr in sorted(by_year.keys()):
        d = by_year[yr]
        ws_cad.cell(row=chart_start, column=1, value=yr)
        ws_cad.cell(row=chart_start, column=2, value=round(d["rev_usd"], 2))
        ws_cad.cell(row=chart_start, column=3, value=round(d["cost_usd"], 2))
        ws_cad.cell(row=chart_start, column=4, value=round(d["rev_cad"], 2))
        ws_cad.cell(row=chart_start, column=5, value=round(d["cost_cad"], 2))
        chart_start += 1

    data_rows = len(by_year)
    chart_data_first_row = chart_start - data_rows  # first data row of chart table
    if data_rows > 0:
        chart_usd = BarChart()
        chart_usd.type = "col"
        chart_usd.title = "REOLUBE46B Duke Energy - Revenue vs Cost (USD)"
        chart_usd.y_axis.title = "Amount (USD)"
        chart_usd.x_axis.title = "Year"
        data = Reference(ws_cad, min_col=2, min_row=chart_data_first_row, max_col=3, max_row=chart_data_first_row + data_rows - 1)
        cats = Reference(ws_cad, min_col=1, min_row=chart_data_first_row, max_row=chart_data_first_row + data_rows - 1)
        chart_usd.add_data(data, titles_from_data=True)
        chart_usd.set_categories(cats)
        chart_usd.width = 14
        chart_usd.height = 8
        ws_usd.add_chart(chart_usd, "M5")

        chart_cad = BarChart()
        chart_cad.type = "col"
        chart_cad.title = "REOLUBE46B Duke Energy - Revenue vs Cost (CAD)"
        chart_cad.y_axis.title = "Amount (CAD)"
        chart_cad.x_axis.title = "Year"
        data2 = Reference(ws_cad, min_col=4, min_row=chart_data_first_row, max_col=5, max_row=chart_data_first_row + data_rows - 1)
        chart_cad.add_data(data2, titles_from_data=True)
        chart_cad.set_categories(cats)
        chart_cad.width = 14
        chart_cad.height = 8
        ws_cad.add_chart(chart_cad, "M5")

    for ws in [ws_usd, ws_cad]:
        for col in range(1, 12):
            ws.column_dimensions[get_column_letter(col)].width = max(10, 14)

    out_dir = os.path.join(os.path.dirname(__file__), "..", "docs")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "REOLUBE46B_Duke_Energy_Orders.xlsx")
    try:
        wb.save(out_path)
    except PermissionError:
        alt = os.path.join(out_dir, "REOLUBE46B_Duke_Energy_Orders_NEW.xlsx")
        wb.save(alt)
        out_path = alt
        print(f"Original file locked - saved to: {alt}")
    print(f"Saved: {out_path}")
    return out_path

if __name__ == "__main__":
    main()
