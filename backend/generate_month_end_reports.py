"""
MONTH-END REPORT GENERATOR
===========================
Generates a comprehensive month-end Excel package combining:
  1. MiSys Inventory on Hand (from Full Company Data CSV export)
  2. Sage Sales Summary for the month (from G Drive Sage CSV export)
  3. Sage Customer Revenue Report
  4. Sage vs MiSys Stock Reconciliation
  5. Best Moving Items (YTD)

Usage:
    python generate_month_end_reports.py [YYYY-MM]
    python generate_month_end_reports.py           # uses current month

Output:
    G:\\Shared drives\\IT_Automation\\MiSys\\Misys Extracted Data\\Daily Snapshots\\YYYY-MM-DD\\
        Month End Package - {Month} {Year}.xlsx

Data sources (all READ-ONLY):
  - MiSys: Full Company Data From Misys\\latest folder\\MIITEM.CSV + MIILOC.CSV
  - Sage:  Full Company Data From SAGE\\latest folder\\tinvent.CSV + tcustomr.CSV + tsalordr.CSV + tsoline.CSV + tinvext.CSV
"""

import os
import sys
import json
import csv
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Configuration ────────────────────────────────────────────────────────────

MISYS_BASE = r"G:\Shared drives\IT_Automation\MiSys\Misys Extracted Data\Full Company Data From Misys"
SAGE_BASE  = r"G:\Shared drives\IT_Automation\MiSys\Misys Extracted Data\Full Company Data From SAGE"
SNAPSHOT_BASE = r"G:\Shared drives\IT_Automation\MiSys\Misys Extracted Data\Daily Snapshots"

MAPPING_FILE = os.path.join(os.path.dirname(__file__), "sage_item_mapping.json")

PRICE_LIST_NAMES = {1: "Regular", 2: "Preferred", 3: "Web Price", 4: "Master"}

# ─── Styling ──────────────────────────────────────────────────────────────────

HDR_DARK  = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
HDR_AMBER = PatternFill(start_color="D97706", end_color="D97706", fill_type="solid")
FILL_GREEN = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
FILL_AMBER = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
FILL_RED   = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
FILL_STRIPE = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

def bold_white(ws, row, col, value, fill=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=True, color="FFFFFF", size=10)
    if fill:
        c.fill = fill
    return c

def header_row(ws, row, cols, fill=HDR_DARK):
    for i, label in enumerate(cols, 1):
        bold_white(ws, row, i, label, fill)

def sf(v, default=0.0):
    if v is None or (isinstance(v, float) and v != v):  # NaN
        return default
    try:
        return float(str(v).strip())
    except Exception:
        return default

def ss(v):
    if v is None:
        return ""
    return str(v).strip()


# ─── Data Loaders ─────────────────────────────────────────────────────────────

def _latest_folder(base: str):
    """Find the most recently modified subfolder in base."""
    if not os.path.isdir(base):
        return None, f"Base folder not found: {base}"
    subs = [f for f in os.listdir(base) if os.path.isdir(os.path.join(base, f)) and not f.startswith('_')]
    if not subs:
        return None, f"No subfolders in {base}"
    subs.sort(key=lambda f: os.path.getmtime(os.path.join(base, f)), reverse=True)
    return os.path.join(base, subs[0]), subs[0]


def _read_csv(folder, name):
    for candidate in [name, name.upper(), name.lower()]:
        path = os.path.join(folder, candidate)
        if os.path.isfile(path):
            with open(path, 'r', encoding='utf-8', errors='replace') as f:
                return list(csv.DictReader(f))
    return []


def load_misys_data():
    folder, label = _latest_folder(MISYS_BASE)
    if not folder:
        print(f"  [MiSys] {label}")
        return [], [], label
    print(f"  [MiSys] Using: {label}")
    miitem = _read_csv(folder, "MIITEM.CSV")
    miiloc = _read_csv(folder, "MIILOC.CSV")
    print(f"  [MiSys] {len(miitem)} items, {len(miiloc)} locations")
    return miitem, miiloc, label


def load_sage_data():
    folder, label = _latest_folder(SAGE_BASE)
    if not folder:
        print(f"  [Sage] {label}")
        return {}, label
    print(f"  [Sage] Using: {label}")
    tables = {}
    for tbl in ["tcustomr", "tinvent", "tinvbyln", "tinvext", "tsalordr", "tsoline", "tinvprc"]:
        rows = _read_csv(folder, f"{tbl}.CSV")
        tables[tbl] = rows
        print(f"  [Sage] {tbl}: {len(rows)} rows")
    return tables, label


def load_item_mapping():
    if not os.path.isfile(MAPPING_FILE):
        return {}
    try:
        with open(MAPPING_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
        result = {}
        for m in data.get('mappings', []):
            if m.get('confirmed') and m.get('sage_part_code'):
                result[ss(m['misys_item_id']).upper()] = m['sage_part_code']
        print(f"  [Mapping] {len(result)} confirmed item mappings")
        return result
    except Exception as e:
        print(f"  [Mapping] Error: {e}")
        return {}


# ─── Sheet builders ───────────────────────────────────────────────────────────

def sheet_misys_inventory(wb, miitem, miiloc, report_month_str):
    ws = wb.create_sheet("MiSys Inventory")
    ws.cell(row=1, column=1, value="CANOIL CANADA LTD.").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Inventory on Hand — {report_month_str}").font = Font(size=12, italic=True)
    ws.cell(row=3, column=1, value="Source: MiSys Full Company Data Export — READ-ONLY").font = Font(size=9, italic=True, color="888888")

    header_row(ws, 5, ["Item No.", "Description", "UOM", "Location(s)", "On Hand", "Unit Cost (Last)", "Value"])

    loc_map: dict = {}
    for r in miiloc:
        iid = ss(r.get('itemId'))
        loc = ss(r.get('locId'))
        if iid and loc:
            loc_map.setdefault(iid, set()).add(loc)

    data_rows = []
    seen = set()
    for item in miitem:
        iid = ss(item.get('itemId'))
        if not iid or iid in seen:
            continue
        seen.add(iid)
        stock = sf(item.get('totQStk'))
        if stock <= 0:
            continue
        cost = sf(item.get('cLast'))
        locs = ', '.join(sorted(loc_map.get(iid, {ss(item.get('locId'))} if item.get('locId') else set())))
        data_rows.append((iid, ss(item.get('descr')), ss(item.get('uOfM')), locs, stock, cost, round(stock * cost, 2)))

    data_rows.sort(key=lambda r: r[0])
    grand_stock = grand_val = 0
    for ri, row in enumerate(data_rows, 6):
        fill = FILL_STRIPE if ri % 2 == 0 else None
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            if fill:
                c.fill = fill
        grand_stock += row[4]
        grand_val += row[6]

    tr = len(data_rows) + 6
    for col, val in [(1, "TOTAL"), (5, round(grand_stock, 4)), (7, round(grand_val, 2))]:
        ws.cell(row=tr, column=col, value=val).font = Font(bold=True)

    for col, w in enumerate([28, 55, 9, 30, 12, 18, 14], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    return len(data_rows), grand_val


def sheet_sage_sales(wb, sage_tables, year, month, report_month_str):
    ws = wb.create_sheet("Sage Sales Summary")
    ws.cell(row=1, column=1, value="SAGE SALES SUMMARY").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"{report_month_str} — READ-ONLY from Sage G Drive Export").font = Font(size=11, italic=True)

    cust_names: dict = {}
    for r in sage_tables.get('tcustomr', []):
        cid = ss(r.get('lId'))
        cust_names[cid] = ss(r.get('sName'))

    inv_names: dict = {}
    for r in sage_tables.get('tinvent', []):
        iid = ss(r.get('lId'))
        inv_names[iid] = (ss(r.get('sPartCode')), ss(r.get('sName')))

    month_prefix = f"{year:04d}-{month:02d}"
    sos = [
        r for r in sage_tables.get('tsalordr', [])
        if ss(r.get('dtSODate', '')).startswith(month_prefix)
        and ss(r.get('bQuote', '0')) != '1'
    ]
    so_ids = {ss(r.get('lId')) for r in sos}

    lines = [r for r in sage_tables.get('tsoline', []) if ss(r.get('lSOId')) in so_ids]

    # Aggregate by customer
    cust_totals: dict = {}
    for so in sos:
        cid = ss(so.get('lCusId'))
        cust_totals.setdefault(cid, {'name': cust_names.get(cid, f'ID:{cid}'), 'total': 0.0, 'order_count': 0})
        cust_totals[cid]['total'] += sf(so.get('dTotal'))
        cust_totals[cid]['order_count'] += 1

    # SO header summary
    ws.cell(row=4, column=1, value=f"Orders this month: {len(sos)}").font = Font(bold=True)
    total_rev = sum(c['total'] for c in cust_totals.values())
    ws.cell(row=5, column=1, value=f"Total Revenue: ${total_rev:,.2f}").font = Font(bold=True, color="166534")

    header_row(ws, 7, ["Customer", "# Orders", "Revenue", "Avg Order Size"], fill=HDR_AMBER)
    sorted_custs = sorted(cust_totals.values(), key=lambda c: c['total'], reverse=True)
    for ri, c in enumerate(sorted_custs, 8):
        fill = FILL_STRIPE if ri % 2 == 0 else None
        for ci, val in enumerate([
            c['name'], c['order_count'],
            round(c['total'], 2),
            round(c['total'] / c['order_count'], 2) if c['order_count'] else 0
        ], 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if fill:
                cell.fill = fill

    # Product breakdown section
    prod_row = len(sorted_custs) + 10
    ws.cell(row=prod_row, column=1, value="Product Sales Breakdown").font = Font(bold=True, size=12)
    header_row(ws, prod_row + 1, ["Part Code", "Description", "Qty Sold", "Revenue"], fill=HDR_AMBER)

    prod_totals: dict = {}
    for line in lines:
        iid = ss(line.get('lInventId'))
        qty = sf(line.get('dQuantity') or line.get('dOrdered') or line.get('dQty') or 0)
        amt = sf(line.get('dAmount') or line.get('dPrice', 0)) * qty
        if iid not in prod_totals:
            part_code, part_name = inv_names.get(iid, (f'ID:{iid}', ''))
            prod_totals[iid] = {'part_code': part_code, 'name': part_name, 'qty': 0.0, 'revenue': 0.0}
        prod_totals[iid]['qty'] += qty
        prod_totals[iid]['revenue'] += sf(line.get('dAmount') or 0)

    sorted_prods = sorted(prod_totals.values(), key=lambda p: p['revenue'], reverse=True)
    for ri, p in enumerate(sorted_prods, prod_row + 2):
        fill = FILL_STRIPE if ri % 2 == 0 else None
        for ci, val in enumerate([p['part_code'], p['name'], round(p['qty'], 2), round(p['revenue'], 2)], 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if fill:
                cell.fill = fill

    for col, w in enumerate([40, 50, 14, 16], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    return len(sos), round(total_rev, 2)


def sheet_customer_revenue(wb, sage_tables, report_month_str):
    ws = wb.create_sheet("Customer Revenue")
    ws.cell(row=1, column=1, value="CUSTOMER REVENUE REPORT").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"YTD — {report_month_str} — READ-ONLY Sage").font = Font(italic=True, size=11)

    header_row(ws, 4, [
        "Customer", "City", "Currency", "Price List",
        "YTD Sales", "Prior Year", "YoY Change %",
        "Credit Limit", "Credit Utilization %",
        "Net Days", "Last Sale"
    ])

    PRICE_LIST_MAP = {1: "Regular", 2: "Preferred", 3: "Web Price", 4: "Master"}
    CURR_MAP = {1: "CAD", 2: "USD"}

    rows = []
    for r in sage_tables.get('tcustomr', []):
        if ss(r.get('bInactive', '0')) == '1':
            continue
        ytd = sf(r.get('dAmtYtd'))
        ly = sf(r.get('dLastYrAmt'))
        credit = sf(r.get('dCrLimit'))
        if ytd <= 0 and ly <= 0:
            continue
        yoy = round((ytd - ly) / ly * 100, 1) if ly > 0 else None
        util = round(ytd / credit * 100, 1) if credit > 0 else None
        try:
            pl = PRICE_LIST_MAP.get(int(sf(r.get('lPrcListId'), 1)), "Regular")
            curr = CURR_MAP.get(int(sf(r.get('lCurrncyId'), 1)), "CAD")
        except Exception:
            pl, curr = "Regular", "CAD"
        rows.append({
            'name': ss(r.get('sName')), 'city': ss(r.get('sCity')), 'currency': curr, 'price_list': pl,
            'ytd': ytd, 'ly': ly, 'yoy': yoy, 'credit': credit, 'util': util,
            'net': int(sf(r.get('nNetDay', 0))), 'last_sale': ss(r.get('dtLastSal')),
        })

    rows.sort(key=lambda r: r['ytd'], reverse=True)
    for ri, r in enumerate(rows, 5):
        fill = FILL_STRIPE if ri % 2 != 0 else None
        util_fill = FILL_RED if (r['util'] or 0) >= 90 else FILL_AMBER if (r['util'] or 0) >= 70 else None
        yoy_fill = FILL_GREEN if (r['yoy'] or 0) > 0 else FILL_RED if (r['yoy'] or 0) < -10 else None

        vals = [r['name'], r['city'], r['currency'], r['price_list'],
                round(r['ytd'], 2), round(r['ly'], 2),
                f"{r['yoy']:+.1f}%" if r['yoy'] is not None else "N/A",
                round(r['credit'], 2),
                f"{r['util']:.1f}%" if r['util'] is not None else "N/A",
                f"Net {r['net']}" if r['net'] else "—", r['last_sale'] or "—"]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if ci == 7 and yoy_fill:
                cell.fill = yoy_fill
            elif ci == 9 and util_fill:
                cell.fill = util_fill
            elif fill:
                cell.fill = fill

    for col, w in enumerate([40, 20, 10, 12, 16, 16, 12, 16, 16, 10, 14], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    return len(rows)


def sheet_best_movers(wb, sage_tables, report_month_str):
    ws = wb.create_sheet("Best Movers")
    ws.cell(row=1, column=1, value="BEST MOVING ITEMS — YTD").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"{report_month_str} — READ-ONLY Sage").font = Font(italic=True, size=11)

    inv_map: dict = {}
    for r in sage_tables.get('tinvent', []):
        iid = ss(r.get('lId'))
        inv_map[iid] = {'part_code': ss(r.get('sPartCode')), 'name': ss(r.get('sName')), 'unit': ss(r.get('sSellUnit'))}

    rows = []
    for r in sage_tables.get('tinvext', []):
        iid = ss(r.get('lInventId'))
        ytd_units = sf(r.get('dYTDUntSld'))
        ytd_rev = sf(r.get('dYTDAmtSld'))
        ly_units = sf(r.get('dPrUntSld'))
        ly_rev = sf(r.get('dPrAmtSld'))
        ytd_cogs = sf(r.get('dYTDCOGS'))
        if ytd_units <= 0 and ly_units <= 0:
            continue
        info = inv_map.get(iid, {'part_code': f'ID:{iid}', 'name': '', 'unit': ''})
        margin = round((ytd_rev - ytd_cogs) / ytd_rev * 100, 1) if ytd_rev > 0 else None
        units_yoy = round((ytd_units - ly_units) / ly_units * 100, 1) if ly_units > 0 else None
        rows.append({
            'part_code': info['part_code'], 'name': info['name'], 'unit': info['unit'],
            'ytd_units': ytd_units, 'ly_units': ly_units, 'units_yoy': units_yoy,
            'ytd_rev': ytd_rev, 'ly_rev': ly_rev, 'ytd_cogs': ytd_cogs,
            'margin': margin, 'last_sold': ss(r.get('dtLastSold')),
        })

    rows.sort(key=lambda r: r['ytd_units'], reverse=True)

    header_row(ws, 4, [
        "#", "Part Code", "Description", "Unit",
        "YTD Units", "Prior Yr Units", "Units YoY %",
        "YTD Revenue", "Prior Yr Revenue",
        "YTD COGS", "Est. Margin %", "Last Sold"
    ])

    for ri, r in enumerate(rows[:50], 5):
        fill = FILL_STRIPE if ri % 2 != 0 else None
        margin_fill = FILL_GREEN if (r['margin'] or 0) >= 30 else FILL_AMBER if (r['margin'] or 0) >= 15 else FILL_RED if r['margin'] is not None else None
        yoy_fill = FILL_GREEN if (r['units_yoy'] or 0) > 0 else FILL_RED if (r['units_yoy'] or 0) < -10 else None
        vals = [
            ri - 4, r['part_code'], r['name'], r['unit'],
            round(r['ytd_units'], 2), round(r['ly_units'], 2),
            f"{r['units_yoy']:+.1f}%" if r['units_yoy'] is not None else "N/A",
            round(r['ytd_rev'], 2), round(r['ly_rev'], 2), round(r['ytd_cogs'], 2),
            f"{r['margin']:.1f}%" if r['margin'] is not None else "N/A",
            r['last_sold'] or "—"
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if ci == 7 and yoy_fill:
                cell.fill = yoy_fill
            elif ci == 11 and margin_fill:
                cell.fill = margin_fill
            elif fill:
                cell.fill = fill

    for col, w in enumerate([5, 28, 45, 8, 12, 14, 11, 16, 16, 14, 13, 14], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    return len(rows)


def sheet_reconciliation(wb, misys_items, sage_tables, item_mapping, report_month_str):
    from daily_inventory_snapshot import add_reconciliation_sheet, sf as _sf
    # Reuse the reconciliation sheet builder from daily_inventory_snapshot.py
    sage_stock_list = []
    inv_map: dict = {}
    for r in sage_tables.get('tinvent', []):
        inv_map[ss(r.get('lId'))] = {'sage_part_code': ss(r.get('sPartCode')), 'sage_name': ss(r.get('sName')), 'sage_in_stock': 0.0, 'sage_last_cost': 0.0, 'sage_cost_of_stock': 0.0}
    for r in sage_tables.get('tinvbyln', []):
        iid = ss(r.get('lInventId'))
        if iid in inv_map:
            inv_map[iid]['sage_in_stock'] += sf(r.get('dInStock'))
            inv_map[iid]['sage_cost_of_stock'] += sf(r.get('dCostStk'))
            lc = sf(r.get('dLastCost'))
            if lc > 0:
                inv_map[iid]['sage_last_cost'] = lc
    sage_stock_list = list(inv_map.values())

    # Convert misys CSV rows to the format expected by add_reconciliation_sheet
    misys_dict = []
    for r in misys_items:
        misys_dict.append({'itemId': ss(r.get('itemId')), 'descr': ss(r.get('descr')),
                           'totQStk': sf(r.get('totQStk')), 'cLast': sf(r.get('cLast'))})

    try:
        match_c, var_c = add_reconciliation_sheet(wb, misys_dict, sage_stock_list, item_mapping)
    except Exception as e:
        ws = wb.create_sheet("Sage vs MiSys Reconciliation")
        ws.cell(row=1, column=1, value=f"Error building reconciliation: {e}")
        match_c, var_c = 0, 0
    return match_c, var_c


# ─── Cover Sheet ──────────────────────────────────────────────────────────────

def sheet_cover(wb, year, month, stats):
    ws = wb.active
    ws.title = "Cover"
    report_month_str = f"{datetime(year, month, 1).strftime('%B %Y')}"

    ws.cell(row=1, column=1, value="CANOIL CANADA LTD.").font = Font(bold=True, size=18)
    ws.cell(row=2, column=1, value=f"Month-End Package — {report_month_str}").font = Font(bold=True, size=14, color="D97706")
    ws.cell(row=3, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(italic=True, size=10, color="888888")
    ws.cell(row=4, column=1, value="All data is READ-ONLY from MiSys and Sage G Drive CSV exports.").font = Font(italic=True, size=9, color="888888")

    ws.cell(row=6, column=1, value="SUMMARY").font = Font(bold=True, size=12)

    for ri, (label, value) in enumerate(stats.items(), 8):
        ws.cell(row=ri, column=1, value=label).font = Font(bold=True)
        ws.cell(row=ri, column=2, value=value)

    ws.cell(row=ri + 2, column=1, value="SHEETS INCLUDED").font = Font(bold=True, size=11)
    sheets = [
        ("MiSys Inventory", "On-hand quantities and values from MiSys"),
        ("Sage Sales Summary", f"Orders placed in {report_month_str} from Sage"),
        ("Customer Revenue", "YTD customer revenue with credit utilization"),
        ("Best Movers", "Top 50 items by YTD units sold"),
        ("Sage vs MiSys Reconciliation", "Stock comparison for confirmed mapped items"),
    ]
    for ri2, (sheet, desc) in enumerate(sheets, ri + 3):
        ws.cell(row=ri2, column=1, value=f"• {sheet}").font = Font(bold=True)
        ws.cell(row=ri2, column=2, value=desc)

    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 50


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    # Parse year/month from args or use current month
    if len(sys.argv) > 1:
        try:
            target = datetime.strptime(sys.argv[1], '%Y-%m')
            year, month = target.year, target.month
        except ValueError:
            print(f"ERROR: Invalid date format '{sys.argv[1]}' — use YYYY-MM")
            sys.exit(1)
    else:
        now = datetime.now()
        year, month = now.year, now.month

    report_month_str = f"{datetime(year, month, 1).strftime('%B %Y')}"
    today = datetime.now().strftime('%Y-%m-%d')

    print("=" * 70)
    print(f"  MONTH-END REPORT GENERATOR — {report_month_str}")
    print(f"  Run date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)

    # Load data
    print("\n  Loading MiSys data...")
    misys_items, misys_locs, misys_label = load_misys_data()

    print("\n  Loading Sage G Drive data...")
    sage_tables, sage_label = load_sage_data()

    print("\n  Loading item mapping...")
    item_mapping = load_item_mapping()

    # Build workbook
    wb = Workbook()

    stats = {}

    print("\n  Building sheets...")

    # 1. MiSys Inventory
    if misys_items:
        count, total_val = sheet_misys_inventory(wb, misys_items, misys_locs, report_month_str)
        stats["MiSys items with stock"] = count
        stats["Total inventory value (MiSys)"] = f"${total_val:,.2f}"
        print(f"    [OK] MiSys Inventory: {count} items, ${total_val:,.2f}")
    else:
        print("    [SKIP] MiSys Inventory: no data")

    # 2. Sage Sales Summary
    if sage_tables.get('tsalordr'):
        so_count, so_rev = sheet_sage_sales(wb, sage_tables, year, month, report_month_str)
        stats[f"Sage SOs in {report_month_str}"] = so_count
        stats[f"Sage revenue {report_month_str}"] = f"${so_rev:,.2f}"
        print(f"    [OK] Sage Sales Summary: {so_count} orders, ${so_rev:,.2f}")
    else:
        print("    [SKIP] Sage Sales Summary: no data")

    # 3. Customer Revenue
    if sage_tables.get('tcustomr'):
        cust_count = sheet_customer_revenue(wb, sage_tables, report_month_str)
        stats["Active Sage customers"] = cust_count
        print(f"    [OK] Customer Revenue: {cust_count} customers")
    else:
        print("    [SKIP] Customer Revenue: no data")

    # 4. Best Movers
    if sage_tables.get('tinvext'):
        mover_count = sheet_best_movers(wb, sage_tables, report_month_str)
        stats["Best movers shown"] = min(mover_count, 50)
        print(f"    [OK] Best Movers: {mover_count} items")
    else:
        print("    [SKIP] Best Movers: no data")

    # 5. Reconciliation
    if misys_items and sage_tables.get('tinvent') and item_mapping:
        match_c, var_c = sheet_reconciliation(wb, misys_items, sage_tables, item_mapping, report_month_str)
        stats["Reconciliation — exact matches"] = match_c
        stats["Reconciliation — variances"] = var_c
        print(f"    [OK] Reconciliation: {match_c} matches, {var_c} variances")
    else:
        print("    [SKIP] Reconciliation: missing data or mappings")

    # Cover sheet (last, uses stats)
    sheet_cover(wb, year, month, stats)

    # Save
    output_dir = os.path.join(SNAPSHOT_BASE, today)
    os.makedirs(output_dir, exist_ok=True)
    filename = f"Month End Package - {report_month_str}.xlsx"
    output_path = os.path.join(output_dir, filename)
    wb.save(output_path)

    print("\n" + "=" * 70)
    print(f"  SAVED: {output_path}")
    print("=" * 70)
    print("\n  Summary:")
    for k, v in stats.items():
        print(f"    {k}: {v}")


if __name__ == "__main__":
    main()
