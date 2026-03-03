"""
DAILY INVENTORY SNAPSHOT
========================
Pulls live inventory data from MISys SQL Server and saves:
  1. Raw data as JSON (permanent archive)
  2. Excel report in the same format used for bank submissions
  3. Sage vs MiSys stock reconciliation (from G Drive Sage CSV export)

Schedule this via Windows Task Scheduler to run once daily (e.g. 6:00 PM).

Data sources:
  - MISys SQL Server (read-only): Host 192.168.1.11 | DB CANOILCA
  - Sage G Drive CSV export (read-only): Full Company Data From SAGE\latest folder\

Output:
  G:\\Shared drives\\IT_Automation\\MiSys\\Misys Extracted Data\\Daily Snapshots\\YYYY-MM-DD\\
    - MIITEM_snapshot.json         (all items with stock/cost data)
    - MIILOC_snapshot.json         (item locations with stock breakdown)
    - Inventory Month End for {date}.xlsx  (ready-to-use Excel, now with Sage reconciliation tab)
    - snapshot_log.txt             (run metadata)
"""

import json
import os
import sys
from datetime import datetime

try:
    import pymssql
except ImportError:
    print("ERROR: pymssql not installed. Run: pip install pymssql")
    sys.exit(1)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# Sage G Drive base path (for reconciliation sheet)
SAGE_GDRIVE_BASE = r"G:\Shared drives\IT_Automation\MiSys\Misys Extracted Data\Full Company Data From SAGE"

# ─── Configuration ───────────────────────────────────────────

MISYS_HOST = '192.168.1.11'
MISYS_DB = 'CANOILCA'
MISYS_USER = 'sa'
MISYS_PASS = 'MISys_SBM1'

SNAPSHOT_BASE = r"G:\Shared drives\IT_Automation\MiSys\Misys Extracted Data\Daily Snapshots"

# ─── SQL Queries ─────────────────────────────────────────────

SQL_MIITEM = """
SELECT
    itemId, descr, xdesc, uOfM, poUOfM, uConvFact,
    type, revId, track, status,
    locId, suplId,
    cLast, cStd, cAvg, cLand,
    totQStk, totQWip, totQRes, totQOrd,
    lstUseDt
FROM MIITEM
ORDER BY itemId
"""

SQL_MIILOC = """
SELECT
    itemId, locId,
    qStk, qWip, qRes, qOrd
FROM MIILOC
ORDER BY itemId, locId
"""

MISYS_TYPES = {
    0: 'Assembled', 1: 'Raw Material', 2: 'Subassembly',
    3: 'Phantom', 4: 'MRP Planned', 5: 'Outside Process',
}


def sf(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).strip())
    except:
        return 0.0


def connect_misys():
    """Connect to MISys SQL Server (read-only)."""
    print("  Connecting to MISys SQL Server...")
    conn = pymssql.connect(
        server=MISYS_HOST,
        user=MISYS_USER,
        password=MISYS_PASS,
        database=MISYS_DB,
        login_timeout=15,
    )
    print("  Connected successfully.")
    return conn


def pull_data(conn):
    """Pull MIITEM and MIILOC from MISys."""
    cursor = conn.cursor(as_dict=True)

    print("  Querying MIITEM...")
    cursor.execute(SQL_MIITEM)
    miitem = cursor.fetchall()
    print(f"    -> {len(miitem)} items")

    print("  Querying MIILOC...")
    cursor.execute(SQL_MIILOC)
    miiloc = cursor.fetchall()
    print(f"    -> {len(miiloc)} location records")

    cursor.close()
    return miitem, miiloc


def save_json(data, path):
    """Save data as JSON with datetime serialization."""
    def handler(obj):
        if isinstance(obj, datetime):
            return obj.isoformat()
        if hasattr(obj, 'Decimal'):
            return float(obj)
        return str(obj)

    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, default=handler, ensure_ascii=False)


def load_sage_stock() -> list:
    """
    Load Sage inventory stock from the latest G Drive CSV export.
    Returns list of dicts: {sage_part_code, sage_name, sage_in_stock, sage_last_cost, sage_cost_of_stock}.
    READ-ONLY — never writes to Sage.
    """
    if not os.path.isdir(SAGE_GDRIVE_BASE):
        print(f"  [SAGE] G Drive base not found: {SAGE_GDRIVE_BASE}")
        return []
    try:
        subfolders = [
            f for f in os.listdir(SAGE_GDRIVE_BASE)
            if os.path.isdir(os.path.join(SAGE_GDRIVE_BASE, f)) and not f.startswith('_')
        ]
        if not subfolders:
            print("  [SAGE] No Sage export subfolders found")
            return []
        subfolders.sort(key=lambda f: os.path.getmtime(os.path.join(SAGE_GDRIVE_BASE, f)), reverse=True)
        folder = os.path.join(SAGE_GDRIVE_BASE, subfolders[0])
        print(f"  [SAGE] Using folder: {subfolders[0]}")
    except Exception as e:
        print(f"  [SAGE] Error listing folders: {e}")
        return []

    try:
        import csv

        # Load tinvent (item master)
        inv_path = os.path.join(folder, "tinvent.CSV")
        if not os.path.isfile(inv_path):
            inv_path = os.path.join(folder, "TINVENT.CSV")
        items: dict = {}
        if os.path.isfile(inv_path):
            with open(inv_path, 'r', encoding='utf-8', errors='replace') as f:
                for row in csv.DictReader(f):
                    iid = row.get('lId', '').strip()
                    if iid:
                        items[iid] = {
                            'sage_part_code': row.get('sPartCode', '').strip(),
                            'sage_name': row.get('sName', '').strip(),
                            'sage_in_stock': 0.0,
                            'sage_last_cost': 0.0,
                            'sage_cost_of_stock': 0.0,
                        }

        # Load tinvbyln (stock on hand)
        byln_path = os.path.join(folder, "tinvbyln.CSV")
        if not os.path.isfile(byln_path):
            byln_path = os.path.join(folder, "TINVBYLN.CSV")
        if os.path.isfile(byln_path):
            with open(byln_path, 'r', encoding='utf-8', errors='replace') as f:
                for row in csv.DictReader(f):
                    iid = row.get('lInventId', '').strip()
                    if iid in items:
                        try:
                            items[iid]['sage_in_stock'] += float(row.get('dInStock') or 0)
                            items[iid]['sage_cost_of_stock'] += float(row.get('dCostStk') or 0)
                            lc = float(row.get('dLastCost') or 0)
                            if lc > 0:
                                items[iid]['sage_last_cost'] = lc
                        except Exception:
                            pass

        result = list(items.values())
        print(f"  [SAGE] Loaded {len(result)} Sage items from G Drive CSV")
        return result
    except Exception as e:
        print(f"  [SAGE] Error loading Sage CSV data: {e}")
        return []


def load_sage_item_mapping() -> dict:
    """
    Load the confirmed MiSys ↔ Sage item mapping from sage_item_mapping.json.
    Returns dict: misys_item_id.upper() -> sage_part_code
    """
    mapping_path = os.path.join(os.path.dirname(__file__), "sage_item_mapping.json")
    if not os.path.isfile(mapping_path):
        return {}
    try:
        with open(mapping_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        result = {}
        for m in data.get('mappings', []):
            if m.get('confirmed') and m.get('sage_part_code'):
                key = str(m['misys_item_id']).strip().upper()
                result[key] = m['sage_part_code']
        print(f"  [MAPPING] Loaded {len(result)} confirmed item mappings")
        return result
    except Exception as e:
        print(f"  [MAPPING] Error loading mapping file: {e}")
        return {}


def add_reconciliation_sheet(wb, miitem, sage_stock_list, item_mapping):
    """Add a Sage vs MiSys reconciliation tab to the workbook."""
    ws = wb.create_sheet("Sage vs MiSys Reconciliation")

    # Header
    ws.cell(row=1, column=1, value="Sage vs MiSys Stock Reconciliation").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(italic=True)
    ws.cell(row=3, column=1, value="Only items with confirmed mappings are compared.").font = Font(italic=True, color="808080")

    headers = ["MiSys Item ID", "MiSys Description", "Sage Part Code", "Sage Description",
               "MiSys Stock", "Sage Stock", "Variance", "MiSys Unit Cost", "Sage Last Cost",
               "Cost Variance", "Status"]
    fill_header = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    fill_ok = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    fill_warn = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    fill_err = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill_header

    # Build Sage lookup: sage_part_code.upper() -> row
    sage_lookup: dict = {}
    for sr in sage_stock_list:
        code = str(sr.get('sage_part_code') or '').strip().upper()
        if code:
            sage_lookup[code] = sr

    # Build MiSys lookup
    misys_lookup: dict = {}
    for item in miitem:
        iid = str(item.get('itemId') or '').strip()
        if iid:
            misys_lookup[iid.upper()] = item

    row_idx = 6
    match_count = 0
    variance_count = 0

    for misys_key_upper, sage_code in sorted(item_mapping.items()):
        sage_code_upper = sage_code.strip().upper()
        misys_item = misys_lookup.get(misys_key_upper)
        sage_item = sage_lookup.get(sage_code_upper)

        misys_stock = sf(misys_item.get('totQStk') if misys_item else 0)
        misys_cost = sf(misys_item.get('cLast') if misys_item else 0)
        misys_desc = str(misys_item.get('descr') or '') if misys_item else ''

        sage_stock = float(sage_item.get('sage_in_stock', 0)) if sage_item else 0.0
        sage_cost = float(sage_item.get('sage_last_cost', 0)) if sage_item else 0.0
        sage_desc = str(sage_item.get('sage_name') or '') if sage_item else ''

        variance = round(misys_stock - sage_stock, 4)
        cost_var = round(misys_cost - sage_cost, 4)

        if not sage_item:
            status = "Sage item not found"
        elif not misys_item:
            status = "MiSys item not found"
        elif abs(variance) < 0.001:
            status = "Match"
            match_count += 1
        elif abs(variance) <= misys_stock * 0.05:
            status = "Minor variance"
            variance_count += 1
        else:
            status = "VARIANCE"
            variance_count += 1

        row_data = [
            misys_key_upper, misys_desc, sage_code, sage_desc,
            misys_stock, sage_stock, variance,
            misys_cost, sage_cost, cost_var, status
        ]

        fill = fill_ok if status in ("Match",) else fill_warn if "Minor" in status else fill_err if "VARIANCE" in status or "not found" in status else None

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            if fill:
                cell.fill = fill

        row_idx += 1

    # Summary
    ws.cell(row=row_idx + 1, column=1, value="SUMMARY").font = Font(bold=True)
    ws.cell(row=row_idx + 2, column=1, value=f"Mapped items: {len(item_mapping)}")
    ws.cell(row=row_idx + 3, column=1, value=f"Exact matches: {match_count}")
    ws.cell(row=row_idx + 4, column=1, value=f"Variances: {variance_count}")

    col_widths = [25, 45, 25, 35, 14, 14, 12, 16, 16, 14, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return match_count, variance_count


def build_report(miitem, miiloc, snapshot_date, output_path, sage_stock_list=None, item_mapping=None):
    """Build the Excel inventory report matching bank format."""

    # Location mapping from MIILOC
    loc_map = {}
    for row in miiloc:
        iid = str(row.get('itemId') or '').strip()
        loc = str(row.get('locId') or '').strip()
        if iid and loc:
            if iid not in loc_map:
                loc_map[iid] = set()
            loc_map[iid].add(loc)

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Month End"

    date_str = snapshot_date.strftime('%B %d, %Y')
    month_str = snapshot_date.strftime('%B %Y')

    ws.cell(row=1, column=3, value="CANOIL CANADA LTD.")
    ws.cell(row=2, column=3, value=f"Inventory Month End \u2013 {month_str}")
    ws.cell(row=3, column=3, value=f"Data as of: {date_str}")

    headers = ["Item No.", "Description", "UOM", "Location", "On Hand", "Unit Cost", "Value"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=5, column=col, value=h).font = Font(bold=True)

    seen = set()
    data_rows = []
    for item in miitem:
        item_no = str(item.get('itemId') or '').strip()
        if not item_no or item_no in seen:
            continue
        seen.add(item_no)

        stock = sf(item.get('totQStk'))
        if stock <= 0:
            continue

        desc = str(item.get('descr') or '').strip()
        uom = str(item.get('uOfM') or '').strip()
        unit_cost = sf(item.get('cLast'))

        if item_no in loc_map:
            location = ', '.join(sorted(loc_map[item_no]))
        else:
            location = str(item.get('locId') or '').strip()

        value = round(stock * unit_cost, 2)
        data_rows.append((item_no, desc, uom, location, stock, unit_cost, value))

    data_rows.sort(key=lambda r: r[0])

    grand_stock = 0
    grand_value = 0
    for ri, row in enumerate(data_rows, 6):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)
        grand_stock += row[4]
        grand_value += row[6]

    total_row = len(data_rows) + 6
    ws.cell(row=total_row, column=4, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=round(grand_stock, 6)).font = Font(bold=True)
    ws.cell(row=total_row, column=7, value=round(grand_value, 2)).font = Font(bold=True)

    for col, w in enumerate([28.7, 59.1, 9.6, 31.0, 12.0, 16.0, 13.0], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Add Sage reconciliation sheet if data is available
    recon_match = 0
    recon_variance = 0
    if sage_stock_list and item_mapping:
        recon_match, recon_variance = add_reconciliation_sheet(wb, miitem, sage_stock_list, item_mapping)

    wb.save(output_path)
    return len(data_rows), grand_stock, grand_value, recon_match, recon_variance


def main():
    now = datetime.now()
    today = now.strftime('%Y-%m-%d')

    print("=" * 65)
    print(f"  DAILY INVENTORY SNAPSHOT - {today}")
    print(f"  Time: {now.strftime('%H:%M:%S')}")
    print("=" * 65)

    # Create output folder
    output_dir = os.path.join(SNAPSHOT_BASE, today)
    os.makedirs(output_dir, exist_ok=True)
    print(f"  Output: {output_dir}")

    try:
        conn = connect_misys()
        miitem, miiloc = pull_data(conn)
        conn.close()
        print("  Connection closed.")
    except Exception as e:
        print(f"\n  ERROR connecting to MISys: {e}")
        print("  Make sure you're on the Canoil network (192.168.1.x)")
        # Write error log
        log_path = os.path.join(output_dir, "snapshot_log.txt")
        with open(log_path, 'w') as f:
            f.write(f"SNAPSHOT FAILED\nDate: {today}\nTime: {now.strftime('%H:%M:%S')}\nError: {e}\n")
        sys.exit(1)

    # Save raw JSON snapshots
    print("\n  Saving raw data...")
    save_json(miitem, os.path.join(output_dir, "MIITEM_snapshot.json"))
    print(f"    -> MIITEM_snapshot.json ({len(miitem)} items)")
    save_json(miiloc, os.path.join(output_dir, "MIILOC_snapshot.json"))
    print(f"    -> MIILOC_snapshot.json ({len(miiloc)} records)")

    # Load Sage data for reconciliation sheet
    print("\n  Loading Sage G Drive CSV data for reconciliation...")
    sage_stock = load_sage_stock()
    item_mapping = load_sage_item_mapping()

    # Generate Excel report
    print("\n  Generating Excel report...")
    date_label = now.strftime('%B %Y')
    xlsx_name = f"Inventory Month End for {date_label}.xlsx"
    xlsx_path = os.path.join(output_dir, xlsx_name)
    items_count, total_stock, total_value, recon_match, recon_variance = build_report(
        miitem, miiloc, now, xlsx_path,
        sage_stock_list=sage_stock if sage_stock else None,
        item_mapping=item_mapping if item_mapping else None
    )
    print(f"    -> {xlsx_name}")
    print(f"       Items with stock: {items_count}")
    print(f"       Total On Hand: {total_stock:,.2f}")
    print(f"       Total Value: ${total_value:,.2f}")
    if sage_stock:
        print(f"       Sage Reconciliation: {recon_match} matches, {recon_variance} variances (tab added)")

    # Write success log
    log_path = os.path.join(output_dir, "snapshot_log.txt")
    with open(log_path, 'w') as f:
        f.write(f"SNAPSHOT SUCCESSFUL\n")
        f.write(f"Date: {today}\n")
        f.write(f"Time: {now.strftime('%H:%M:%S')}\n")
        f.write(f"MIITEM records: {len(miitem)}\n")
        f.write(f"MIILOC records: {len(miiloc)}\n")
        f.write(f"Items with stock > 0: {items_count}\n")
        f.write(f"Total On Hand: {total_stock:,.6f}\n")
        f.write(f"Total Value: ${total_value:,.2f}\n")
        f.write(f"Sage items loaded: {len(sage_stock)}\n")
        f.write(f"Confirmed item mappings: {len(item_mapping)}\n")
        f.write(f"Reconciliation: {recon_match} exact matches, {recon_variance} variances\n")
        f.write(f"Output: {output_dir}\n")

    print("\n" + "=" * 65)
    print("  SNAPSHOT COMPLETE")
    print("=" * 65)


if __name__ == "__main__":
    main()
