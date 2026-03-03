import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\Haron\Downloads\Inventory Month End for October 2025.xlsx', data_only=True)
ws = wb['Stock Transfers']

print(f'Total rows: {ws.max_row}')
print(f'Total cols: {ws.max_column}')

# Check last 20 rows for totals/summary
print('\n--- LAST 20 ROWS ---')
for row_idx in range(max(1, ws.max_row - 20), ws.max_row + 1):
    parts = []
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        v = cell.value
        if v is not None:
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            bold = ' BOLD' if cell.font and cell.font.bold else ''
            parts.append(f'[{col_letter}]{bold} {repr(v)[:60]}')
    if parts:
        print(f'  Row {row_idx}: {"  |  ".join(parts)}')

# Check unique categories
categories = set()
locations = set()
date_range = [None, None]
total_cost_out = 0
total_cost_in = 0
for row_idx in range(2, ws.max_row + 1):
    cat = ws.cell(row=row_idx, column=2).value
    loc = ws.cell(row=row_idx, column=5).value
    dt = ws.cell(row=row_idx, column=1).value
    cost_out = ws.cell(row=row_idx, column=6).value or 0
    cost_in = ws.cell(row=row_idx, column=7).value or 0
    if cat:
        categories.add(cat)
    if loc:
        locations.add(loc)
    if dt:
        if date_range[0] is None or dt < date_range[0]:
            date_range[0] = dt
        if date_range[1] is None or dt > date_range[1]:
            date_range[1] = dt
    total_cost_out += float(cost_out) if cost_out else 0
    total_cost_in += float(cost_in) if cost_in else 0

print(f'\nUnique categories: {sorted(categories)}')
print(f'Unique locations: {sorted(locations)}')
print(f'Date range: {date_range[0]} to {date_range[1]}')
print(f'Total rows (data): {ws.max_row - 1}')
print(f'Total Cost Out: ${total_cost_out:,.5f}')
print(f'Total Cost In: ${total_cost_in:,.5f}')
