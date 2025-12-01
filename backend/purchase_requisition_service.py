"""
Purchase Requisition Service
Auto-fills requisition forms from PO data with Excel generation
"""

from flask import Blueprint, jsonify, request, send_file
import os
import json
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import zipfile
import shutil

pr_service = Blueprint('pr_service', __name__)

# Use environment variable for path (works for both local and Cloud Run)
# Default to Windows path for local, but can be overridden via environment variable
GDRIVE_BASE = os.getenv('GDRIVE_BASE', r"G:\Shared drives\IT_Automation\MiSys\Misys Extracted Data\API Extractions")
# Use relative path for Docker compatibility
_current_dir = os.path.dirname(os.path.abspath(__file__))
PR_TEMPLATE = os.path.join(_current_dir, 'templates', 'purchase_requisition', 'PR-2025-06-09-Lanxess.xlsx')


def preserve_template_structure(template_path, output_bytes):
    """
    openpyxl corrupts many parts of xlsx files (VML, EMF images, printer settings, etc).
    This function rebuilds the output by:
    1. Starting with the complete original template
    2. Only replacing the worksheet XML with the modified version from openpyxl
    
    This preserves ALL template structure including drawings, signatures, logos, etc.
    """
    try:
        output_bytes.seek(0)
        
        # Files that openpyxl modifies and we WANT to keep from openpyxl output
        # (the actual cell data changes)
        files_from_openpyxl = {
            'xl/worksheets/sheet1.xml',  # Main worksheet with our data
            'xl/worksheets/sheet2.xml',  # Instructions sheet (if exists)
        }
        
        new_output = io.BytesIO()
        
        with zipfile.ZipFile(template_path, 'r') as template_zip:
            with zipfile.ZipFile(output_bytes, 'r') as saved_zip:
                with zipfile.ZipFile(new_output, 'w', zipfile.ZIP_DEFLATED) as new_zip:
                    # Start with ALL files from original template
                    for name in template_zip.namelist():
                        if name in files_from_openpyxl and name in saved_zip.namelist():
                            # Use openpyxl's version (has our cell changes)
                            new_zip.writestr(name, saved_zip.read(name))
                        else:
                            # Use original template version (preserves drawings, etc.)
                            new_zip.writestr(name, template_zip.read(name))
        
        new_output.seek(0)
        return new_output
        
    except Exception as e:
        print(f"Warning: Could not preserve template structure: {e}")
        import traceback
        traceback.print_exc()
        output_bytes.seek(0)
        return output_bytes


# Cache for latest folder to avoid repeated API calls
_latest_folder_cache = None
_latest_folder_cache_time = None

def get_latest_folder():
    """
    Get latest MISys API extraction folder
    Always returns the most recent folder by date (YYYY-MM-DD format)
    Works for both local and Google Cloud Run environments
    
    Caches result for 5 minutes to avoid repeated slow API calls
    """
    global _latest_folder_cache, _latest_folder_cache_time
    
    # Check cache (valid for 5 minutes)
    if _latest_folder_cache and _latest_folder_cache_time:
        cache_age = (datetime.now() - _latest_folder_cache_time).total_seconds()
        if cache_age < 300:  # 5 minutes
            return _latest_folder_cache
    
    try:
        # On Cloud Run, use Google Drive API
        is_cloud_run = os.getenv('K_SERVICE') is not None
        
        if is_cloud_run:
            # Import dynamically to avoid circular imports
            import sys
            if 'app' in sys.modules:
                app_module = sys.modules['app']
                get_google_drive_service = getattr(app_module, 'get_google_drive_service', None)
                
                if get_google_drive_service:
                    print("SEARCH: Using Google Drive API to find latest folder (from PR service)...")
                    service = get_google_drive_service()
                    if service:
                        latest_id, latest_name = service.find_latest_api_extractions_folder()
                        if latest_name:
                            print(f"✅ Latest MISys API extraction folder (via API): {latest_name}")
                            _latest_folder_cache = latest_name
                            _latest_folder_cache_time = datetime.now()
                            return latest_name
            return None
        
        # Local environment - read from G: drive
        if not os.path.exists(GDRIVE_BASE):
            print(f"⚠️ GDRIVE_BASE path not accessible: {GDRIVE_BASE}")
            return None
        
        # Get all folders
        folders = [f for f in os.listdir(GDRIVE_BASE) if os.path.isdir(os.path.join(GDRIVE_BASE, f))]
        
        if not folders:
            print(f"⚠️ No folders found in: {GDRIVE_BASE}")
            return None
        
        # Sort by folder name (assuming YYYY-MM-DD format) - most recent first
        folders.sort(reverse=True)
        latest_folder = folders[0]
        
        print(f"✅ Latest MISys API extraction folder: {latest_folder}")
        _latest_folder_cache = latest_folder
        _latest_folder_cache_time = datetime.now()
        return latest_folder
        
    except Exception as e:
        print(f"❌ Error getting latest folder: {e}")
        import traceback
        traceback.print_exc()
        return None


def load_items():
    """Load items from latest folder"""
    latest = get_latest_folder()
    if not latest:
        return []
    
    items_file = os.path.join(GDRIVE_BASE, latest, 'Items.json')
    if os.path.exists(items_file):
        with open(items_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []


def load_po_data(po_number):
    """Load merged PO data"""
    latest = get_latest_folder()
    if not latest:
        return None
    
    merged_folder = os.path.join(GDRIVE_BASE, latest, 'MERGED_POS', 'individual_pos')
    po_file = os.path.join(merged_folder, f'PO_{po_number}.json')
    
    if os.path.exists(po_file):
        with open(po_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    return None


def get_inventory_data(item_no):
    """Get current inventory data for an item from CustomAlert5.json
    
    On Cloud Run, this returns None since local G: drive is not accessible.
    The frontend passes stock data from the already-loaded inventory data.
    """
    try:
        # Check if we're on Cloud Run - local G: drive not accessible
        is_cloud_run = os.getenv('K_SERVICE') is not None
        if is_cloud_run:
            # On Cloud Run, skip local file reads - frontend provides stock data
            return None
        
        # Local environment - read from G: drive
        if not os.path.exists(GDRIVE_BASE):
            return None
            
        latest = get_latest_folder()
        if not latest:
            return None
            
        folder_path = os.path.join(GDRIVE_BASE, latest)
        custom_alert_file = os.path.join(folder_path, 'CustomAlert5.json')
        
        if not os.path.exists(custom_alert_file):
            return None
        
        with open(custom_alert_file, 'r', encoding='utf-8') as f:
            inventory_data = json.load(f)
        
        # Find the item in inventory data
        for item in inventory_data:
            if item.get('Item No.') == item_no:
                return {
                    'stock': float(item.get('Stock', 0)),
                    'wip': float(item.get('WIP', 0)),
                    'reserve': float(item.get('Reserve', 0)),
                    'on_order': float(item.get('On Order', 0)),
                    'minimum': float(item.get('Minimum', 0)),
                    'maximum': float(item.get('Maximum', 0)),
                    'reorder_level': float(item.get('Reorder Level', 0)),
                    'reorder_quantity': float(item.get('Reorder Quantity', 0)),
                    'recent_cost': float(item.get('Recent Cost', 0).replace('$', '').replace(',', '')) if item.get('Recent Cost') else 0,
                    'average_cost': float(item.get('Average Cost', 0).replace('$', '').replace(',', '')) if item.get('Average Cost') else 0,
                    'landed_cost': float(item.get('Landed Cost', 0).replace('$', '').replace(',', '')) if item.get('Landed Cost') else 0,
                    'stocking_units': item.get('Stocking Units', ''),
                    'purchasing_units': item.get('Purchasing Units', ''),
                    'units_conversion_factor': float(item.get('Units Conversion Factor', 1).replace(',', '')) if item.get('Units Conversion Factor') else 1
                }
        
        return None
        
    except Exception as e:
        print(f"Error getting inventory data for {item_no}: {e}")
        return None

def get_recent_purchase_price(item_no, limit=5):
    """
    Get recent purchase prices for an item from PO details
    Returns list of recent purchases sorted by date (most recent first)
    """
    try:
        latest = get_latest_folder()
        if not latest:
            return []
        
        folder_path = os.path.join(GDRIVE_BASE, latest)
        po_details_file = os.path.join(folder_path, 'PurchaseOrderDetails.json')
        po_file = os.path.join(folder_path, 'PurchaseOrders.json')
        
        if not os.path.exists(po_details_file) or not os.path.exists(po_file):
            return []
        
        with open(po_details_file, 'r', encoding='utf-8') as f:
            po_details = json.load(f)
        
        with open(po_file, 'r', encoding='utf-8') as f:
            pos = json.load(f)
        
        # Create a lookup for PO headers by PO number
        po_lookup = {po.get('PO No.'): po for po in pos}
        
        # Find all purchases for this item
        item_purchases = []
        print(f"[DEBUG] Looking for item: {item_no}")
        print(f"[DEBUG] Total PO details records: {len(po_details)}")
        print(f"[DEBUG] Total PO headers: {len(pos)}")
        
        for detail in po_details:
            if detail.get('Item No.') == item_no:
                po_no = detail.get('PO No.')
                po_header = po_lookup.get(po_no, {})
                
                print(f"[DEBUG] Found item {item_no} in PO {po_no}")
                print(f"[DEBUG] Unit Price: {detail.get('Unit Price', 0)}")
                print(f"[DEBUG] Supplier: {po_header.get('Name', '')}")
                
                purchase = {
                    'item_no': item_no,  # Include the item number
                    'po_no': po_no,
                    'unit_price': detail.get('Unit Price', 0),
                    'order_date': po_header.get('Order Date', ''),  # From PO header
                    'supplier_no': po_header.get('Supplier No.', ''),  # From PO header
                    'quantity': detail.get('Ordered', 0),
                    'total_amount': round(detail.get('Unit Price', 0) * detail.get('Ordered', 0), 2),
                    'description': detail.get('Description', ''),
                    'supplier_name': po_header.get('Name', ''),
                    'contact': po_header.get('Contact', ''),
                    'terms': po_header.get('Terms', ''),
                    'purchase_unit': detail.get('Purchase U/M', ''),  # Real purchase unit
                    'supplier_item_no': detail.get('Supplier Item No.', '')  # Supplier's item number
                }
                item_purchases.append(purchase)
        
        # Sort by order date (most recent first)
        item_purchases.sort(key=lambda x: x.get('order_date', ''), reverse=True)
        
        # Return the most recent purchase (limit=1 for latest price)
        return item_purchases[:1] if item_purchases else []
        
    except Exception as e:
        print(f"Error getting recent prices for {item_no}: {e}")
        return []


def get_supplier_info(supplier_no):
    """
    Get supplier information from most recent PO
    First tries merged PO data (more complete), then falls back to PurchaseOrders.json
    """
    try:
        latest = get_latest_folder()
        if not latest:
            return None
        
        folder_path = os.path.join(GDRIVE_BASE, latest)
        
        # FIRST: Try to get supplier info from merged PO data (more complete)
        merged_folder = os.path.join(folder_path, 'MERGED_POS', 'individual_pos')
        if os.path.exists(merged_folder):
            # Find all merged PO files for this supplier
            merged_pos = []
            for filename in os.listdir(merged_folder):
                if filename.startswith('PO_') and filename.endswith('.json'):
                    try:
                        po_file = os.path.join(merged_folder, filename)
                        with open(po_file, 'r', encoding='utf-8') as f:
                            merged_po = json.load(f)
                        
                        # Check if this PO is for the supplier we're looking for
                        supplier = merged_po.get('Supplier', {})
                        if supplier.get('Supplier_No') == supplier_no:
                            merged_pos.append(merged_po)
                    except Exception as e:
                        print(f"Error reading merged PO file {filename}: {e}")
                        continue
            
            if merged_pos:
                # Sort by Order Date (most recent first)
                merged_pos.sort(key=lambda x: x.get('Order_Info', {}).get('Order_Date', ''), reverse=True)
                most_recent_po = merged_pos[0]
                
                supplier = most_recent_po.get('Supplier', {})
                order_info = most_recent_po.get('Order_Info', {})
                financial = most_recent_po.get('Financial', {})
                
                print(f"✅ Found supplier info from merged PO data: {supplier.get('Name', '')}")
                
                return {
                    'supplier_no': supplier_no,
                    'name': supplier.get('Name', ''),
                    'contact': supplier.get('Contact', ''),
                    'phone': supplier.get('Phone', '') or supplier.get('Phone_No', '') or supplier.get('Telephone', ''),
                    'email': supplier.get('Email', ''),
                    'terms': order_info.get('Terms', ''),
                    'po_count': len(merged_pos),
                    'last_order_date': order_info.get('Order_Date', ''),
                    'last_po_no': most_recent_po.get('PO_Number', ''),
                    'buyer': order_info.get('Buyer', ''),
                    'currency': financial.get('Currency', {}).get('Home_Currency', ''),
                    'total_amount': financial.get('Total_Amount', 0)
                }
        
        # FALLBACK: Use PurchaseOrders.json if merged data not available
        po_file = os.path.join(folder_path, 'PurchaseOrders.json')
        
        if not os.path.exists(po_file):
            return None
        
        with open(po_file, 'r', encoding='utf-8') as f:
            pos = json.load(f)
        
        # Find most recent PO for this supplier
        supplier_pos = [po for po in pos if po.get('Supplier No.') == supplier_no]
        
        if not supplier_pos:
            return None
        
        # Sort by Order Date (most recent first)
        supplier_pos.sort(key=lambda x: x.get('Order Date', ''), reverse=True)
        most_recent_po = supplier_pos[0]
        
        print(f"✅ Found supplier info from PurchaseOrders.json: {most_recent_po.get('Name', '')}")
        
        return {
            'supplier_no': supplier_no,
            'name': most_recent_po.get('Name', ''),
            'contact': most_recent_po.get('Contact', ''),
            'phone': most_recent_po.get('Phone', ''),  # This field might not exist
            'email': most_recent_po.get('Email', ''),  # This field might not exist
            'terms': most_recent_po.get('Terms', ''),
            'po_count': len(supplier_pos),
            'last_order_date': most_recent_po.get('Order Date', ''),
            'last_po_no': most_recent_po.get('PO No.', ''),
            'buyer': most_recent_po.get('Buyer', ''),
            'currency': most_recent_po.get('Home Currency', ''),
            'total_amount': most_recent_po.get('Total Amount', 0)
        }
        
    except Exception as e:
        print(f"Error getting supplier info for {supplier_no}: {e}")
        import traceback
        traceback.print_exc()
        return None


@pr_service.route('/api/pr/search-items', methods=['GET'])
def search_items():
    """Smart search for items with predictions"""
    try:
        query = request.args.get('q', '').lower()
        limit = int(request.args.get('limit', 20))
        
        if not query or len(query) < 2:
            return jsonify({"results": []})
        
        items = load_items()
        matches = []
        
        for item in items:
            item_no = str(item.get('Item No.', '')).lower()
            description = str(item.get('Description', '')).lower()
            
            # Score matches
            score = 0
            if query in item_no:
                score += 100
            if query in description:
                score += 50
            if item_no.startswith(query):
                score += 200
            
            if score > 0:
                matches.append({
                    'score': score,
                    'item_no': item.get('Item No.'),
                    'description': item.get('Description'),
                    'recent_cost': item.get('Recent Cost', 0),
                    'standard_cost': item.get('Standard Cost', 0),
                    'stocking_units': item.get('Stocking Units', 'EA'),
                    'purchasing_units': item.get('Purchasing Units', 'EA'),
                    'preferred_supplier': item.get('Preferred Supplier Number', ''),
                    'reorder_quantity': item.get('Reorder Quantity', 0),
                    'minimum': item.get('Minimum', 0),
                    'reorder_level': item.get('Reorder Level', 0)
                })
        
        # Sort by score and limit
        matches.sort(key=lambda x: x['score'], reverse=True)
        results = matches[:limit]
        
        # Remove score from results
        for r in results:
            del r['score']
        
        return jsonify({"results": results})
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@pr_service.route('/api/pr/item-details/<item_no>', methods=['GET'])
def get_item_details(item_no):
    """Get complete details for an item"""
    try:
        items = load_items()
        item = next((i for i in items if i.get('Item No.') == item_no), None)
        
        if not item:
            return jsonify({"error": "Item not found"}), 404
        
        return jsonify({
            "item_no": item.get('Item No.'),
            "description": item.get('Description'),
            "extended_description": item.get('Extended Description', ''),
            "recent_cost": item.get('Recent Cost', 0),
            "standard_cost": item.get('Standard Cost', 0),
            "average_cost": item.get('Average Cost', 0),
            "landed_cost": item.get('Landed Cost', 0),
            "stocking_units": item.get('Stocking Units', 'EA'),
            "purchasing_units": item.get('Purchasing Units', 'EA'),
            "preferred_supplier": item.get('Preferred Supplier Number', ''),
            "preferred_location": item.get('Preferred Location Number', ''),
            "minimum": item.get('Minimum', 0),
            "maximum": item.get('Maximum', 0),
            "reorder_level": item.get('Reorder Level', 0),
            "reorder_quantity": item.get('Reorder Quantity', 0)
        })
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@pr_service.route('/api/pr/item-pricing/<item_no>', methods=['GET'])
def get_item_pricing(item_no):
    """Get recent purchase prices for an item"""
    try:
        recent_prices = get_recent_purchase_price(item_no, 5)
        
        if not recent_prices:
            return jsonify({
                "item_no": item_no,
                "recent_prices": [],
                "message": "No recent purchase history found"
            })
        
        return jsonify({
            "item_no": item_no,
            "recent_prices": recent_prices,
            "latest_price": recent_prices[0]['unit_price'] if recent_prices else 0
        })
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@pr_service.route('/api/pr/item-enhanced/<item_no>', methods=['GET'])
def get_item_enhanced(item_no):
    """Get item details with recent pricing and supplier information"""
    try:
        items = load_items()
        item = next((i for i in items if i.get('Item No.') == item_no), None)
        
        if not item:
            return jsonify({"error": "Item not found"}), 404
        
        # Get recent purchase prices
        recent_prices = get_recent_purchase_price(item_no, 3)
        
        # Get supplier info if preferred supplier exists
        supplier_info = None
        preferred_supplier = item.get('Preferred Supplier Number', '')
        if preferred_supplier:
            supplier_info = get_supplier_info(preferred_supplier)
        
        return jsonify({
            "item_no": item.get('Item No.'),
            "description": item.get('Description'),
            "recent_cost": item.get('Recent Cost', 0),
            "standard_cost": item.get('Standard Cost', 0),
            "preferred_supplier": preferred_supplier,
            "recent_prices": recent_prices,
            "supplier_info": supplier_info,
            "reorder_quantity": item.get('Reorder Quantity', 0),
            "minimum": item.get('Minimum', 0),
            "reorder_level": item.get('Reorder Level', 0),
            "stocking_units": item.get('Stocking Units', 'EA'),
            "purchasing_units": item.get('Purchasing Units', 'EA')
        })
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@pr_service.route('/api/pr/generate', methods=['POST'])
def generate_requisition():
    """Generate filled Purchase Requisition Excel file"""
    try:
        data = request.get_json()
        
        # Required fields
        user_info = data.get('user_info', {})
        items = data.get('items', [])
        supplier = data.get('supplier', {})
        
        # DEBUG: Print what we received
        print("\n" + "="*80)
        print("🔍 DEBUG: RECEIVED REQUEST DATA")
        print("="*80)
        print(f"User Info: {user_info}")
        print(f"Items Count: {len(items)}")
        print(f"Supplier Data: {supplier}")
        print(f"Supplier Name: {supplier.get('name', 'EMPTY')}")
        print(f"Supplier Contact: {supplier.get('contact', 'EMPTY')}")
        print(f"Supplier Address: {supplier.get('address', {})}")
        print("="*80 + "\n")
        
        if not items:
            return jsonify({"error": "No items provided"}), 400
        
        # Load template
        if not os.path.exists(PR_TEMPLATE):
            return jsonify({"error": "Template not found"}), 404
        
        # Load template - DO NOT use keep_vba=True as it corrupts non-VBA xlsx files
        wb = openpyxl.load_workbook(PR_TEMPLATE, data_only=False)
        
        sheet = wb['Purchase Requisition']
        
        # DO NOT CLEAR CELLS - This breaks merged cells and form structure
        # Only fill the specific cells needed, preserving all formatting and structure
        # For item rows we're not using, we'll leave them as-is (template may have formulas/formatting)
        
        # Fill user info (CORRECT CELLS FROM TEMPLATE)
        sheet['I7'] = user_info.get('name', 'Haron Alhakimi')  # I7: REQUESTED BY
        sheet['I9'] = user_info.get('department', 'Sales')  # I9: DEPARTMENT
        sheet['B5'] = user_info.get('justification', 'Low Stock')  # B5: JUSTIFICATION
        sheet['B13'] = user_info.get('lead_time', '4 weeks')  # B13: LEAD TIME
        
        # Fill dates
        today = datetime.now()
        sheet['I11'] = today.strftime('%Y-%m-%d')  # I11: DATE REQUESTED (drives formula in I5)
        
        # Calculate date needed (add lead time weeks)
        try:
            lead_time_weeks = int(user_info.get('lead_time', '4').split()[0])
        except:
            lead_time_weeks = 4
        date_needed = today + timedelta(weeks=lead_time_weeks)
        sheet['I13'] = date_needed.strftime('%Y-%m-%d')  # I13: DATE NEEDED
        
        # AUTO-DETECT SUPPLIER INFO FROM FIRST ITEM'S RECENT PURCHASE DATA
        auto_supplier_info = None
        if items:
            first_item = items[0]
            item_no = first_item.get('item_no', '')
            
            # Get recent purchase data for this item
            recent_prices = get_recent_purchase_price(item_no, limit=1)
            
            if recent_prices:
                latest_purchase = recent_prices[0]
                supplier_no = latest_purchase.get('supplier_no', '')
                supplier_name = latest_purchase.get('supplier_name', '')
                contact = latest_purchase.get('contact', '')
                terms = latest_purchase.get('terms', '')
                
                # Get full supplier info
                supplier_info = get_supplier_info(supplier_no)
                
                if supplier_info:
                    auto_supplier_info = {
                        'name': supplier_name,
                        'contact': contact,
                        'email': supplier_info.get('email', ''),
                        'phone': supplier_info.get('phone', ''),
                        'terms': f"{terms} days" if terms else "",
                        'currency': supplier_info.get('currency', 'CAD'),
                        'buyer': supplier_info.get('buyer', ''),
                        'po_count': supplier_info.get('po_count', 0)
                    }
                    
                    print(f"✅ AUTO-DETECTED SUPPLIER: {supplier_name}")
                    print(f"   Contact: {contact}")
                    print(f"   Terms: {auto_supplier_info['terms']}")
                    print(f"   Buyer: {auto_supplier_info['buyer']}")
                    print(f"   Total POs: {auto_supplier_info['po_count']}")
                else:
                    print(f"⚠️ No full supplier info found for {supplier_no}")
            else:
                print(f"⚠️ No recent purchase data found for item: {item_no}")
        
        # Fill supplier info - use auto-detected if available, otherwise use manual input
        if auto_supplier_info:
            supplier_name = auto_supplier_info.get('name', '')
            supplier_contact = auto_supplier_info.get('contact', '')
            supplier_email = auto_supplier_info.get('email', '')
            supplier_phone = auto_supplier_info.get('phone', '')
            supplier_terms = auto_supplier_info.get('terms', '')
        else:
            supplier_name = supplier.get('name', '')
            supplier_contact = supplier.get('contact', '')
            supplier_email = supplier.get('email', '')
            supplier_phone = supplier.get('phone', '')
            supplier_terms = supplier.get('terms', '')
        
        # === VENDOR INFO (Based on actual template layout) ===
        # B9: Vendor Name
        if supplier_name:
            sheet['B9'] = supplier_name
        
        # B11: Contact Name (with phone if available)
        if supplier_contact:
            contact_with_phone = supplier_contact
            if supplier_phone:
                contact_with_phone += f" | {supplier_phone}"
            sheet['B11'] = contact_with_phone
        elif supplier_phone:
            sheet['B11'] = supplier_phone
        
        # === VENDOR ADDRESS (Column C) ===
        address = supplier.get('address', {})
        
        # Build full address string
        address_parts = []
        address_line_1 = ''
        full_address = ''
        
        if address:
            address_line_1 = address.get('Line_1', '') or address.get('Address', '')
            city = address.get('City', '')
            state = address.get('State', '')
            postal_code = address.get('Postal_Code', '')
            country = address.get('Country', '')
            
            if address_line_1:
                address_parts.append(address_line_1)
            
            # City, State, Postal
            city_state_zip = []
            if city:
                city_state_zip.append(city)
            if state:
                city_state_zip.append(state)
            if postal_code:
                city_state_zip.append(postal_code)
            if city_state_zip:
                address_parts.append(', '.join(city_state_zip))
            
            if country:
                address_parts.append(country)
            
            full_address = '\n'.join(address_parts) if address_parts else ''
        
        # C9: Full Address (multi-line)
        if full_address:
            sheet['C9'] = full_address
        elif supplier_name:
            sheet['C9'] = 'Address not available'
        
        # C11: Email
        if supplier_email:
            sheet['C11'] = supplier_email
        elif supplier_terms:
            sheet['C11'] = f"Terms: {supplier_terms}"
        
        print("\n" + "="*60)
        print(f"[PR] Filling {len(items)} items into Excel...")
        print(f"[PR] User: {user_info.get('name')} | Dept: {user_info.get('department')}")
        print(f"[PR] Supplier: {supplier_name} | Contact: {supplier_contact}")
        print("="*60)
        
        # Fill line items - starting at row 16 (first data row after headers at row 15)
        start_row = 16
        for idx, item in enumerate(items):
            if idx >= 10:  # Limit to 10 items
                break
            
            row = start_row + idx
            
            # Clear only this specific row's data cells before filling (preserves form structure)
            for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
                sheet[f'{col}{row}'].value = ''  # Clear only data cells for this row
            
            item_no = item.get('item_no', '')
            description = item.get('description', '')
            current_stock = item.get('current_stock', '')
            unit = item.get('unit', 'EA')
            quantity = item.get('quantity', 0)
            unit_price = item.get('unit_price', 0)
            
            # Get real pricing, supplier data, and inventory data for this specific item
            item_pricing_data = None
            inventory_data = None
            days_since_last_order = None
            if item_no:
                # Get recent pricing data
                recent_prices = get_recent_purchase_price(item_no, limit=1)
                if recent_prices:
                    item_pricing_data = recent_prices[0]
                    # Use real pricing data if available
                    if not unit_price or unit_price == 0:
                        unit_price = item_pricing_data.get('unit_price', 0)
                    if not unit or unit == 'EA':
                        unit = item_pricing_data.get('purchase_unit', unit)
                    
                    # Calculate days since last order
                    order_date_str = item_pricing_data.get('order_date', '')
                    if order_date_str:
                        try:
                            # Parse date (format: YYYY-MM-DD or similar)
                            from datetime import datetime
                            order_date = datetime.strptime(order_date_str[:10], '%Y-%m-%d')
                            days_since_last_order = (datetime.now() - order_date).days
                            item_pricing_data['days_since_last_order'] = days_since_last_order
                        except Exception as e:
                            print(f"[PR] Could not parse order date '{order_date_str}': {e}")
                
                # Get real inventory data
                inventory_data = get_inventory_data(item_no)
                if inventory_data:
                    # Use real stock data (Quantity On Hand) - override the input current_stock
                    current_stock = inventory_data.get('stock', 0)
                    print(f"[PR] Using real inventory data (Qty On Hand): Stock = {current_stock}")
                    # Use inventory pricing if no PO pricing available
                    if not unit_price or unit_price == 0:
                        unit_price = inventory_data.get('recent_cost', 0)
                    if not unit or unit == 'EA':
                        unit = inventory_data.get('purchasing_units', unit)
                else:
                    print(f"[PR] No inventory data found for {item_no} - using provided stock: {current_stock}")
            
            print(f"[PR] Row {row}: {item_no} | Desc: {description[:30]} | Qty: {quantity} | Price: ${unit_price} | Unit: {unit} | Stock: {current_stock}")
            
            sheet[f'B{row}'] = item_no           # B: ITEM № (NOT A!)
            sheet[f'C{row}'] = description       # C: PRODUCT/SERVICE DESCRIPTION
            
            # D: INVENTORY TURNOVER (IN DAYS) - days since last order
            inventory_turnover = item.get('inventory_turnover', '')
            if inventory_turnover:
                sheet[f'D{row}'] = inventory_turnover
            elif item_pricing_data and item_pricing_data.get('days_since_last_order'):
                sheet[f'D{row}'] = item_pricing_data.get('days_since_last_order')
            
            sheet[f'E{row}'] = current_stock     # E: CURRENT STOCK (from real inventory data)
            sheet[f'F{row}'] = unit              # F: PURCHASING UNIT (from real data)
            sheet[f'G{row}'] = int(quantity) if quantity else 0     # G: QTY (as number)
            sheet[f'H{row}'] = round(float(unit_price), 2) if unit_price else 0.0  # H: UNIT PRICE (from real data, rounded to 2 decimals)
            # I: TOTAL column has formula =G17*H17 already in template
        
        print(f"[PR] Completed filling {min(len(items), 10)} items")
        print("="*60 + "\n")
        
        # Fix formulas for item rows (should be =G16*H16, etc.)
        # DO NOT modify row heights - preserve original form structure
        num_items = len(items[:10])
        for idx in range(num_items):
            row = 16 + idx
            sheet[f'I{row}'] = f'=G{row}*H{row}'
            # DO NOT modify row heights - this can break signature areas and form structure
        
        # Clear unused item rows (rows beyond what we filled) - only clear data cells for those specific rows
        # Use empty strings instead of None to preserve cell structure and merged cells
        for row in range(16 + num_items, 26):
            for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
                # Set to empty string - this preserves merged cell structures better than None
                sheet[f'{col}{row}'].value = ''
        
        # Update TOTAL row formula to sum all item totals (I30 = SUM of I16:I25 or however many items)
        last_item_row = 16 + num_items - 1
        sheet['I30'] = f'=SUM(I16:I{last_item_row})'
        print(f"[PR] Total formula set: =SUM(I16:I{last_item_row})")
        
        # DO NOT modify alignment, row heights, or formatting
        # This preserves the original form structure including signature areas
        # The template already has proper formatting - we only fill values
        
        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Rebuild file using template structure (openpyxl corrupts drawings/media)
        output = preserve_template_structure(PR_TEMPLATE, output)
        
        # Generate filename
        filename = f"PR_{today.strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Error generating requisition: {e}")
        return jsonify({"error": str(e)}), 500


@pr_service.route('/api/pr/from-po/<po_number>', methods=['POST'])
def generate_from_po(po_number):
    """Generate requisition from existing PO data"""
    try:
        data = request.get_json()
        user_info = data.get('user_info', {})
        selected_items = data.get('items', [])  # Items with quantities updated by user
        
        # Load PO data
        po = load_po_data(po_number)
        if not po:
            return jsonify({"error": f"PO {po_number} not found"}), 404
        
        # Prepare supplier info from PO (merged PO data structure)
        supplier = {
            'name': po.get('Supplier', {}).get('Name', ''),
            'contact': po.get('Supplier', {}).get('Contact', ''),
            'email': po.get('Supplier', {}).get('Email', ''),
            'phone': po.get('Supplier', {}).get('Phone', '') or po.get('Supplier', {}).get('Phone_No', '') or po.get('Supplier', {}).get('Telephone', ''),
            'address': po.get('Shipping_Billing', {}).get('Ship_To_Address', {}) or po.get('Shipping_Billing', {}).get('Bill_To_Address', {})
        }
        
        # Also try to get supplier info from merged PO data using supplier number
        supplier_no = po.get('Supplier', {}).get('Supplier_No', '')
        if supplier_no:
            supplier_info = get_supplier_info(supplier_no)
            if supplier_info:
                # Enhance supplier info with data from get_supplier_info
                supplier['email'] = supplier.get('email') or supplier_info.get('email', '')
                supplier['phone'] = supplier.get('phone') or supplier_info.get('phone', '')
                supplier['terms'] = supplier_info.get('terms', '')
        
        print(f"[PR FROM PO] Supplier: {supplier.get('name')} | Contact: {supplier.get('contact')}")
        
        # Prepare items - use user-selected items or all from PO
        items_to_add = []
        if selected_items:
            # User provided specific items with quantities
            for sel_item in selected_items:
                items_to_add.append({
                    'item_no': sel_item['item_no'],
                    'description': sel_item['description'],
                    'unit': sel_item.get('unit', 'EA'),
                    'quantity': sel_item['quantity'],
                    'unit_price': sel_item.get('unit_price', 0),
                    'current_stock': sel_item.get('current_stock', ''),
                    'inventory_turnover': sel_item.get('inventory_turnover', '')
                })
        else:
            # Use all items from PO
            for item in po.get('Line_Items', []):
                unit_price = item.get('Item_Master', {}).get('Cost_History', {}).get('Recent_Cost', 0)
                if unit_price == 0:
                    unit_price = item['Pricing']['Unit_Cost']
                
                items_to_add.append({
                    'item_no': item['Item_No'],
                    'description': item['Description'],
                    'unit': item['Pricing']['Purchase_Unit_of_Measure'],
                    'quantity': item['Pricing']['Quantity_Ordered'],
                    'unit_price': unit_price,
                    'current_stock': '',
                    'inventory_turnover': ''
                })
        
        # Generate requisition
        return generate_requisition_internal(user_info, items_to_add, supplier)
        
    except Exception as e:
        print(f"Error generating from PO: {e}")
        return jsonify({"error": str(e)}), 500


def generate_requisition_internal(user_info, items, supplier):
    """
    Internal function to generate requisition
    Uses the same logic as generate_requisition() but accepts pre-prepared data
    """
    try:
        # Load template with all formatting preserved
        if not os.path.exists(PR_TEMPLATE):
            return jsonify({"error": "Template not found"}), 404
        
        try:
            # DO NOT use keep_vba=True as it corrupts non-VBA xlsx files
            wb = openpyxl.load_workbook(PR_TEMPLATE, data_only=False)
        except Exception as e:
            print(f"Error loading template: {e}")
            return jsonify({"error": f"Failed to load template: {str(e)}"}), 500
        
        sheet = wb['Purchase Requisition']
        
        # Fill user info (CORRECT CELLS FROM TEMPLATE - same as main function)
        sheet['I7'] = user_info.get('name', 'Haron Alhakimi')  # I7: REQUESTED BY
        sheet['I9'] = user_info.get('department', 'Sales')  # I9: DEPARTMENT
        sheet['B5'] = user_info.get('justification', 'Low Stock')  # B5: JUSTIFICATION
        sheet['B13'] = user_info.get('lead_time', '4 weeks')  # B13: LEAD TIME
        
        # Fill dates
        today = datetime.now()
        sheet['I11'] = today.strftime('%Y-%m-%d')  # I11: DATE REQUESTED
        
        # Calculate date needed (add lead time weeks)
        try:
            lead_time_weeks = int(user_info.get('lead_time', '4').split()[0])
        except:
            lead_time_weeks = 4
        date_needed = today + timedelta(weeks=lead_time_weeks)
        sheet['I13'] = date_needed.strftime('%Y-%m-%d')  # I13: DATE NEEDED
        
        # Fill supplier info (CORRECT CELLS FROM TEMPLATE)
        supplier_name = supplier.get('name', '')
        supplier_contact = supplier.get('contact', '')
        supplier_email = supplier.get('email', '')
        supplier_phone = supplier.get('phone', '')
        supplier_terms = supplier.get('terms', '')
        
        # === VENDOR INFO (Based on actual template layout) ===
        # B9: Vendor Name
        if supplier_name:
            sheet['B9'] = supplier_name
        
        # B11: Contact Name (with phone if available)
        if supplier_contact:
            contact_with_phone = supplier_contact
            if supplier_phone:
                contact_with_phone += f" | {supplier_phone}"
            sheet['B11'] = contact_with_phone
        elif supplier_phone:
            sheet['B11'] = supplier_phone
        
        # === VENDOR ADDRESS (Column C) ===
        address = supplier.get('address', {})
        
        # Build full address string
        address_parts = []
        
        if address:
            address_line_1 = address.get('Line_1', '') or address.get('Address', '')
            city = address.get('City', '')
            state = address.get('State', '') or address.get('Province', '')
            postal_code = address.get('Postal_Code', '') or address.get('PostalCode', '')
            country = address.get('Country', '')
            
            if address_line_1:
                address_parts.append(address_line_1)
            
            # City, State, Postal
            city_state_zip = []
            if city:
                city_state_zip.append(city)
            if state:
                city_state_zip.append(state)
            if postal_code:
                city_state_zip.append(postal_code)
            if city_state_zip:
                address_parts.append(', '.join(city_state_zip))
            
            if country:
                address_parts.append(country)
        
        # C9: Full Address (multi-line)
        if address_parts:
            sheet['C9'] = '\n'.join(address_parts)
        elif supplier_name:
            sheet['C9'] = 'Address not available'
        
        # C11: Email
        if supplier_email:
            sheet['C11'] = supplier_email
        elif supplier_terms:
            sheet['C11'] = f"Terms: {supplier_terms}"
        
        # Fill line items - starting at row 16 (same as main function)
        start_row = 16
        for idx, item in enumerate(items):
            if idx >= 10:  # Limit to 10 items
                break
            
            row = start_row + idx
            
            # Clear only this specific row's data cells before filling
            for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
                sheet[f'{col}{row}'].value = ''
            
            item_no = item.get('item_no', '')
            description = item.get('description', '')
            current_stock = item.get('current_stock', '')
            unit = item.get('unit', 'EA')
            quantity = item.get('quantity', 0)
            unit_price = item.get('unit_price', 0)
            
            # Get real pricing and inventory data if item_no is available
            item_pricing_data = None
            inventory_data = None
            if item_no:
                recent_prices = get_recent_purchase_price(item_no, limit=1)
                if recent_prices:
                    item_pricing_data = recent_prices[0]
                    if not unit_price or unit_price == 0:
                        unit_price = item_pricing_data.get('unit_price', 0)
                    if not unit or unit == 'EA':
                        unit = item_pricing_data.get('purchase_unit', unit)
                
                inventory_data = get_inventory_data(item_no)
                if inventory_data:
                    current_stock = inventory_data.get('stock', 0)
                    print(f"[PR INTERNAL] Using real inventory data (Qty On Hand): Stock = {current_stock}")
                    if not unit_price or unit_price == 0:
                        unit_price = inventory_data.get('recent_cost', 0)
                    if not unit or unit == 'EA':
                        unit = inventory_data.get('purchasing_units', unit)
            
            sheet[f'B{row}'] = item_no
            sheet[f'C{row}'] = description
            
            # D: Add supplier item number if available
            if item_pricing_data and item_pricing_data.get('supplier_item_no'):
                sheet[f'D{row}'] = item_pricing_data.get('supplier_item_no', '')
            
            sheet[f'E{row}'] = current_stock
            sheet[f'F{row}'] = unit
            sheet[f'G{row}'] = int(quantity) if quantity else 0
            sheet[f'H{row}'] = round(float(unit_price), 2) if unit_price else 0.0
        
        # Fix formulas for item rows
        num_items = len(items[:10])
        for idx in range(num_items):
            row = 16 + idx
            sheet[f'I{row}'] = f'=G{row}*H{row}'
        
        # Clear unused item rows
        for row in range(16 + num_items, 26):
            for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
                sheet[f'{col}{row}'].value = ''
        
        # Update TOTAL row formula
        last_item_row = 16 + num_items - 1
        if num_items > 0:
            sheet['I30'] = f'=SUM(I16:I{last_item_row})'
        
        # Save to memory
        output = io.BytesIO()
        try:
            wb.save(output)
            output.seek(0)
            
            # Verify the file was saved correctly
            if output.getvalue() is None or len(output.getvalue()) == 0:
                raise Exception("Workbook save resulted in empty file")
            
            # Restore VML drawings (signature shapes) that openpyxl doesn't preserve
            output = preserve_vml_drawings(PR_TEMPLATE, output)
            
            filename = f"PR_{today.strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
        except Exception as e:
            print(f"Error saving workbook: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({"error": f"Failed to save Excel file: {str(e)}"}), 500
        finally:
            # Ensure workbook is closed
            try:
                wb.close()
            except:
                pass
                
    except Exception as e:
        print(f"Error in generate_requisition_internal: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Failed to generate requisition: {str(e)}"}), 500


@pr_service.route('/api/pr/supplier/<supplier_no>', methods=['GET'])
def get_supplier_details(supplier_no):
    """Get full supplier details from most recent PO"""
    try:
        # Use helper function to get robust supplier info
        supplier_info = get_supplier_info(supplier_no)
        
        if not supplier_info:
            # Fallback if helper fails (e.g. invalid supplier number)
            return jsonify({
                'supplier_no': supplier_no,
                'error': 'Supplier not found in PO history',
                'note': 'Please enter supplier details manually'
            }), 404
            
        # Add note about address availability
        supplier_info['note'] = 'Supplier address not available in system - please enter manually'
        supplier_info['address'] = {} # Explicitly set empty address so frontend knows to prompt
        
        return jsonify(supplier_info)
        
    except Exception as e:
        print(f"Error getting supplier details: {e}")
        return jsonify({"error": str(e)}), 500

