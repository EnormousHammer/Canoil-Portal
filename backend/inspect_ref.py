import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\Haron\Downloads\Inventory Month End for October 2025.xlsx', data_only=True)
print('Sheet names:', wb.sheetnames)

for sname in wb.sheetnames:
    ws = wb[sname]
    print(f'\n{"="*60}')
    print(f'SHEET: {sname}')
    print(f'Dimensions: {ws.dimensions}')
    print(f'Rows: {ws.max_row}, Cols: {ws.max_column}')
    merged = list(ws.merged_cells.ranges)
    if merged:
        print(f'Merged cells: {merged}')
    print(f'{"="*60}')

    for row_idx in range(1, min(50, ws.max_row + 1)):
        parts = []
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            v = cell.value
            if v is not None:
                fmt = cell.number_format if cell.number_format != 'General' else ''
                font_info = ''
                if cell.font and cell.font.bold:
                    font_info = ' BOLD'
                fill_info = ''
                if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb and cell.fill.start_color.rgb != '00000000':
                    fill_info = f' bg={cell.fill.start_color.rgb}'
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                val_str = repr(v)[:80]
                parts.append(f'[{col_letter}{row_idx}]{font_info}{fill_info} {val_str} fmt={fmt}')
        if parts:
            print(f'  Row {row_idx}:')
            for p in parts:
                print(f'    {p}')

    # Also show column widths
    print(f'\n  Column widths:')
    for col_idx in range(1, ws.max_column + 1):
        cl = openpyxl.utils.get_column_letter(col_idx)
        w = ws.column_dimensions[cl].width
        if w:
            print(f'    {cl}: {w}')
